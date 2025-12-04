
import os
import math
from django.shortcuts import render
from django.http import JsonResponse
from .models import FlightData,FdmFlightData
from datetime import date, datetime
from .models import *
from django.contrib.auth.decorators import login_required
from .utils import format_file_size

@login_required(login_url='login')
def dashboard_view(request):
    query = request.GET.get('query', '')
    selected_date = request.GET.get('date', '')

    # Set filter date based on selected date or default to today
    filter_date = date.today() if not selected_date else datetime.strptime(selected_date, "%Y-%m-%d").date()

    # Check if it's an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Filter based on query and selected date, then order by std_utc
        schedules = (
            FlightData.objects.filter(sd_date_utc=filter_date).filter(
                flight_no__icontains=query
            ) | FlightData.objects.filter(
                sd_date_utc=filter_date
            ).filter(
                dep_code_iata__icontains=query
            ) | FlightData.objects.filter(
                sd_date_utc=filter_date
            ).filter(
                arr_code_iata__icontains=query
            ) | FlightData.objects.filter(
                sd_date_utc=filter_date
            ).filter(
                dep_code_icao__icontains=query
            ) | FlightData.objects.filter(
                sd_date_utc=filter_date
            ).filter(
                arr_code_icao__icontains=query
            ) | FlightData.objects.filter(
                sd_date_utc=filter_date
            ).filter(
                tail_no__icontains=query)

        ).order_by('std_utc')

        # Serialize data for JSON response
        data = list(schedules.values('sd_date_utc', 'flight_no', 'tail_no','dep_code_iata', 'dep_code_icao',
                                      'arr_code_iata', 'arr_code_icao', 'std_utc', 'atd_utc',
                                      'takeoff_utc', 'touchdown_utc', 'ata_utc', 'sta_utc'))
        return JsonResponse(data, safe=False)
    
    # Non-AJAX request loads all today's flights, ordered by std_utc
    schedules = FlightData.objects.filter(sd_date_utc=filter_date).order_by('std_utc')
    return render(request, 'aimsintegration/dashboard.html', {'schedules': schedules})


#Cargo Project API

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils.dateparse import parse_date
from .models import FlightData, CargoFlightData
from .serializers import FlightDataSerializer
from datetime import datetime
from django.utils.decorators import method_decorator

@method_decorator(login_required(login_url='login'), name='dispatch')
class FlightDataListView(APIView):
    def get(self, request, *args, **kwargs):
        # Extract query parameters
        origin_icao = request.query_params.get('origin_icao')
        destination_icao = request.query_params.get('destination_icao')
        scheduled_departure_date = request.query_params.get('scheduled_departure_date', None)

        # If no scheduled_departure_date is provided, use today's date
        if scheduled_departure_date:
            try:
                scheduled_departure_date = parse_date(scheduled_departure_date)
                if not scheduled_departure_date:
                    raise ValueError("Invalid date format")
            except ValueError:
                return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            scheduled_departure_date = datetime.today().date()

        # Query the database based on provided filters
        flight_data = CargoFlightData.objects.filter(
            dep_code_icao=origin_icao,
            arr_code_icao=destination_icao,
            sd_date_utc__gte=scheduled_departure_date
        )

        # Serialize the data using the updated serializer
        serializer = FlightDataSerializer(flight_data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



#CPAT Project API
from django.http import JsonResponse
from django.shortcuts import render
from django.utils.timezone import now
from .models import CompletionRecord
from datetime import datetime
from dateutil.relativedelta import relativedelta
import calendar

# Validity periods for courses
VALIDITY_PERIODS = {
    "FRMS": 12,     # Fatigue Education & Awareness Training
    "ETPG": 36,     # ETOPS Ground
    "LVO-G": 36,    # LVO Ground    
    "PBNGRN": 12,   # PBN Ground
    "RVSMGS": 0,    # RVSM Ground (never expires)
    "A330C1": 6,
    "FAS": 12,
    "B737C1": 6,
    "Q400C1": 6,
    "TCAS": 36,
    "ADW": 36,
    "PWS": 36,
}

def calculate_expiry_date(completion_date_str, course_code):
    """
    Calculate expiry date based on completion date in DDMMYYYY format
    and adjust to the last day of the expiry month.
    """
    if not completion_date_str:
        return "--"  # No completion date available

    try:
        # Parse the input date in YYYY-MM-DD format
        completion_date = datetime.strptime(completion_date_str, "%Y-%m-%d")
        validity_period = VALIDITY_PERIODS.get(course_code, 0)

        if validity_period == 0:
            return "--"  # No expiry date (never expires)

        # Calculate the tentative expiry date
        expiry_date = completion_date + relativedelta(months=validity_period)
        # Adjust to the last day of the expiry month
        last_day = calendar.monthrange(expiry_date.year, expiry_date.month)[1]
        expiry_date = expiry_date.replace(day=last_day)
        return expiry_date.strftime("%Y-%m-%d")  # Return as YYYY-MM-DD
    except ValueError:
        return "--"

@login_required(login_url='login')
def todays_completion_records_view(request):
    today         = now().date()
    query         = request.GET.get('query', '').strip()
    selected_date = request.GET.get('date', '').strip()

    # 1) Base queryset + filters
    qs = CompletionRecord.objects.all()

    if selected_date:
        try:
            d  = datetime.strptime(selected_date, "%Y-%m-%d").date()
            qs = qs.filter(completion_date=d)
        except ValueError:
            qs = qs.none()

    if query:
        qs = (
            qs.filter(employee_id__icontains=query) |
            qs.filter(employee_email__icontains=query) |
            qs.filter(course_code__icontains=query)
        )

    qs = qs.order_by("-completion_date")

    # 2) Paginate
    paginator   = Paginator(qs, 10)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # 3) AJAX → JSON for just this page
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        data = []
        for rec in page_obj:
            comp_str = rec.completion_date.strftime("%Y-%m-%d") if rec.completion_date else ""
            data.append({
                "id":              rec.id,
                "employee_id":     rec.employee_id,
                "employee_email":  rec.employee_email,
                "course_code":     rec.course_code,
                "completion_date": comp_str,
                "expiry_date":     calculate_expiry_date(comp_str, rec.course_code),
                "validity_period": VALIDITY_PERIODS.get(rec.course_code, "--"),
                "score":           rec.score or "--",
                "time_in_seconds": rec.time_in_seconds or "--",
                "start_date":      rec.start_date.strftime("%Y-%m-%d") if rec.start_date else "--",
                "end_date":        rec.end_date.strftime("%Y-%m-%d")   if rec.end_date   else "--",
            })
        return JsonResponse(data, safe=False)

    # 4) Non-AJAX → enrich & render
    enriched = []
    for rec in page_obj:
        comp_str = rec.completion_date.strftime("%Y-%m-%d") if rec.completion_date else ""
        rec.expiry_date     = calculate_expiry_date(comp_str, rec.course_code)
        rec.validity_period = VALIDITY_PERIODS.get(rec.course_code, "--")
        enriched.append(rec)

    display_pages = get_display_pages(page_obj, num_links=2)

    return render(request, "aimsintegration/cpat_completion_records.html", {
        "records":       enriched,
        "page_obj":      page_obj,
        "display_pages": display_pages,
        "today":         today,
        "query":         query,
        "selected_date": selected_date,
    })



#FDM

from django.http import JsonResponse
from django.shortcuts import render
from .models import FdmFlightData
import logging
import datetime as dt
from django.utils.timezone import make_aware


logger = logging.getLogger(__name__)

@login_required(login_url='login')
def fdm_dashboard_view(request):
    query = request.GET.get('query', '').strip()  # Get the query string, remove leading/trailing spaces
    selected_date = request.GET.get('date', '').strip()  # Get the selected date, remove spaces

    # Set the filter date
    if selected_date:
        try:
            filter_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        except ValueError:
            logger.error(f"Invalid date format received: {selected_date}")
            return JsonResponse({"error": "Invalid date format. Use 'YYYY-MM-DD'."}, status=400)
    else:
        filter_date = date.today()

    # Convert filter_date to a timezone-aware datetime object (UTC)
    # filter_date = make_aware(datetime.combine(filter_date, datetime.min.time()), timezone=timezone.utc)
    
    filter_date = make_aware(dt.datetime.combine(filter_date, dt.datetime.min.time()),dt.timezone.utc)

    logger.info(f"Selected Date: {selected_date}, Filter Date (UTC): {filter_date}")

    # Check if it's an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Build the filter for the query
        filters = (
            FdmFlightData.objects.filter(sd_date_utc=filter_date).filter(flight_no__icontains=query) |
            FdmFlightData.objects.filter(sd_date_utc=filter_date).filter(dep_code_iata__icontains=query) |
            FdmFlightData.objects.filter(sd_date_utc=filter_date).filter(arr_code_iata__icontains=query) |
            FdmFlightData.objects.filter(sd_date_utc=filter_date).filter(dep_code_icao__icontains=query) |
            FdmFlightData.objects.filter(sd_date_utc=filter_date).filter(arr_code_icao__icontains=query) |
            FdmFlightData.objects.filter(sd_date_utc=filter_date).filter(tail_no__icontains=query)
        ).distinct().order_by('std_utc')

        # Serialize data for JSON response
        data = list(filters.values(
            'sd_date_utc', 'flight_no', 'tail_no', 'dep_code_icao', 'arr_code_icao',
            'std_utc', 'atd_utc', 'takeoff_utc', 'touchdown_utc', 'ata_utc',
            'sta_utc', 'flight_type', 'etd_utc', 'eta_utc'
        ))

        logger.info(f"Filtered schedules (AJAX): {len(data)} records found.")
        return JsonResponse(data, safe=False)

    # Handle normal (non-AJAX) requests
    fdm_schedules = FdmFlightData.objects.filter(sd_date_utc=filter_date).order_by('std_utc')
    logger.info(f"Filtered schedules (non-AJAX): {len(fdm_schedules)} records found.")

    return render(request, 'aimsintegration/fdm.html', {'fdm_schedules': fdm_schedules})




from django.http import JsonResponse
from .models import CrewMember
from datetime import datetime
import logging
from django.db.models import Case, When, Value, IntegerField
logger = logging.getLogger(__name__)

@login_required(login_url='login')
def get_crew_details(request):
    # Get parameters from the frontend
    flight_no = request.GET.get('flight_no')
    origin = request.GET.get('origin')
    destination = request.GET.get('destination')
    date = request.GET.get('date')  # Expected as 'YYYY-MM-DD'

    logger.info(f"Fetching crew details for flight_no={flight_no}, origin={origin}, destination={destination}, sd_date_utc={date}")

    # Validate and parse the date
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        logger.error(f"Invalid date format received: {date}")
        return JsonResponse({"error": "Invalid date format. Use 'YYYY-MM-DD'."}, status=400)

    # Fetch all matching crew members
    # crew_members = CrewMember.objects.filter(
    #     flight_no=flight_no,
    #     origin=origin,
    #     destination=destination,
    #     sd_date_utc=date_obj
    # ).values('name', 'role', 'flight_no', 'origin', 'destination', 'crew_id')
    

    crew_members = CrewMember.objects.filter(
        flight_no=flight_no,
        origin=origin,
        destination=destination,
        sd_date_utc=date_obj
    ).values('name', 'role', 'flight_no', 'origin', 'destination', 'crew_id').order_by(
        Case(
            When(role='CP', then=Value(1)),  # Assign priority 1 to CP (Captain)
            default=Value(2),  # Assign priority 2 to all other roles
            output_field=IntegerField()
        )
    )


    # Log the query results
    logger.info(f"Crew members found: {list(crew_members)}")

    # Return the response
    return JsonResponse(list(crew_members), safe=False)





# Crew Allowance project
import csv
import io
from datetime import datetime, date
from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Sum
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

import openpyxl

from .models import Crew, Duty, Airport, Invoice, InvoiceItem
from .forms import CSVUploadForm

# @login_required(login_url='login')
def get_display_pages(page_obj, num_links=2):
    """
    Returns a list of page numbers around the current page,
    plus '...' if there's a gap, plus the very first and last pages if needed.
    """
    current_page = page_obj.number
    total_pages = page_obj.paginator.num_pages

    # If total_pages is small, just show all pages.
    if total_pages <= num_links * 2 + 6:
        return list(range(1, total_pages + 1))

    # Otherwise, build a “window” around current_page.
    pages = []
    left_bound = current_page - num_links
    right_bound = current_page + num_links

    # Always show page #1
    pages.append(1)

    # "..." if there's a gap between 1 and left_bound
    if left_bound > 2:
        pages.append("...")

    # Pages in [left_bound..right_bound], clipped to [2..(total_pages-1)]
    for p in range(max(left_bound, 2), min(right_bound, total_pages - 1) + 1):
        pages.append(p)

    # "..." if there's a gap between right_bound and total_pages - 1
    if right_bound < total_pages - 1:
        pages.append("...")

    # Always show the last page
    pages.append(total_pages)

    return pages


@login_required(login_url='login')
def upload_callowance_file(request):
    """
    Allows user to upload a CSV or TXT file that may contain multiple months of data.
    """
    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['file']
            # Parse and generate invoices for all months found
            handle_callowance_csv(csv_file)  # Removed login_required from this function
            return redirect('crew_allowance_list')
    else:
        form = CSVUploadForm()

    return render(request, 'aimsintegration/callowance_upload.html', {
        'form': form,
    })

# Optimized version of handle_callowance_csv
from django.db import transaction

def handle_callowance_csv(csv_file):
    """
    1. Reads CSV/TXT rows.
    2. Converts layover hh:mm -> minutes in Duty.
    3. Creates/updates Crew, Duty, etc.
    4. For each distinct (Crew, Month) in the file, automatically creates/updates
       an Invoice and its InvoiceItems. That way if the CSV has multiple months,
       each month gets its own Invoice.
    
    Optimized for performance with batch processing.
    """
    decoded_file = csv_file.read().decode('utf-8', errors='replace').splitlines()
    reader = csv.reader(decoded_file)

    # We'll use a transaction for better performance
    with transaction.atomic():
        # First pass: collect all data to minimize DB hits
        duties_by_crew_month = {}
        
        # Collection dictionaries
        airport_codes = set()        # All IATA codes we need
        crews_data = {}              # crew_id -> data
        duties_to_create = []        # List of duties to bulk create
        
        # Determine date format once, not per row
        date_formats = ["%d/%m/%y", "%Y-%m-%d"]
        
        for idx, row in enumerate(reader, start=1):
            # Safe extraction of each column; adjust indices to match your CSV
            date_str    = row[0].strip() if len(row) > 0 else ""
            crew_id_str = row[1].strip() if len(row) > 1 else ""
            first_name  = row[2].strip() if len(row) > 2 else ""
            last_name   = row[3].strip() if len(row) > 3 else ""
            position    = row[4].strip() if len(row) > 4 else ""
            flight_no   = row[5].strip() if len(row) > 5 else ""
            tail_no     = row[6].strip() if len(row) > 6 else ""
            dep_code    = row[7].strip().upper() if len(row) > 7 else ""
            arr_code    = row[8].strip().upper() if len(row) > 8 else ""
            layover_str = row[9].strip() if len(row) > 9 else ""

            # Parse the date 
            duty_date = None
            for fmt in date_formats:
                try:
                    duty_date = datetime.strptime(date_str, fmt).date()
                    break
                except ValueError:
                    pass

            if not duty_date:
                # If we can't parse the date, skip this row
                continue

            # Parse layover hh:mm => total minutes
            layover_minutes = 0
            if ":" in layover_str:
                try:
                    hh, mm = layover_str.split(":")
                    layover_minutes = int(hh)*60 + int(mm)
                except ValueError:
                    pass

            # Store crew info for batch processing
            if not crew_id_str:
                crew_id_str = f"no_id_{idx}"
                
            crews_data[crew_id_str] = {
                'first_name': first_name,
                'last_name': last_name,
                'position': position
            }
            
            # Collect airport codes
            if dep_code:
                airport_codes.add(dep_code)
            if arr_code:
                airport_codes.add(arr_code)
                
            # Store duty info for later creation
            duties_to_create.append({
                'crew_id': crew_id_str,
                'duty_date': duty_date,
                'flight_no': flight_no,
                'tail_no': tail_no,
                'dep_code': dep_code,
                'arr_code': arr_code,
                'layover_minutes': layover_minutes,
                'month_first': date(duty_date.year, duty_date.month, 1)
            })
        
        # Process all airport codes in one go
        airport_map = {}  # iata_code -> Airport object
        if airport_codes:
            # Get existing airports
            for airport in Airport.objects.filter(iata_code__in=airport_codes):
                airport_map[airport.iata_code] = airport
                
            # Create missing airports in bulk
            airports_to_create = []
            for code in airport_codes:
                if code not in airport_map:
                    airports_to_create.append(Airport(iata_code=code))
                    
            if airports_to_create:
                created_airports = Airport.objects.bulk_create(airports_to_create)
                for airport in created_airports:
                    airport_map[airport.iata_code] = airport
        
        # Process all crews in one go
        crew_map = {}  # crew_id -> Crew object
        if crews_data:
            # Get existing crews
            for crew in Crew.objects.filter(crew_id__in=crews_data.keys()):
                crew_map[crew.crew_id] = crew
                # Update existing crew data
                crew_data = crews_data[crew.crew_id]
                if (crew.first_name != crew_data['first_name'] or 
                    crew.last_name != crew_data['last_name'] or 
                    crew.position != crew_data['position']):
                    crew.first_name = crew_data['first_name']
                    crew.last_name = crew_data['last_name']
                    crew.position = crew_data['position']
                    crew.save()
                    
            # Create missing crews in bulk
            crews_to_create = []
            for crew_id, crew_data in crews_data.items():
                if crew_id not in crew_map:
                    crews_to_create.append(Crew(
                        crew_id=crew_id, 
                        first_name=crew_data['first_name'],
                        last_name=crew_data['last_name'],
                        position=crew_data['position']
                    ))
                    
            if crews_to_create:
                created_crews = Crew.objects.bulk_create(crews_to_create)
                for crew in created_crews:
                    crew_map[crew.crew_id] = crew
        
        # Process duties and organize by crew & month
        duties_batch = []
        duty_id_map = {}  # Map to track duties by (crew_id, month)
        
        for duty_data in duties_to_create:
            crew = crew_map[duty_data['crew_id']]
            dep_airport = airport_map.get(duty_data['dep_code'])
            arr_airport = airport_map.get(duty_data['arr_code'])
            
            # Create Duty objects
            duty = Duty(
                duty_date=duty_data['duty_date'],
                crew=crew,
                flight_number=duty_data['flight_no'],
                departure_airport=dep_airport,
                arrival_airport=arr_airport,
                layover_time_minutes=duty_data['layover_minutes'],
                tail_number=duty_data['tail_no']
            )
            duties_batch.append(duty)
            
            # Track duties by crew & month for invoice generation
            month_first = duty_data['month_first']
            key = (crew.id, month_first)
            if key not in duty_id_map:
                duty_id_map[key] = []
            
            # Add to tracking (will update with real IDs after save)
            duty_id_map[key].append(duty)
            
            # Bulk create in batches of 500 for better performance
            if len(duties_batch) >= 500:
                created_duties = Duty.objects.bulk_create(duties_batch)
                # Clear batch for next round
                duties_batch = []
        
        # Create any remaining duties
        if duties_batch:
            created_duties = Duty.objects.bulk_create(duties_batch)
        
        # Now generate Invoices and InvoiceItems
        for (crew_id, month_day), duty_list in duty_id_map.items():
            # Get the crew
            cr = Crew.objects.get(id=crew_id)
            
            # Create or get invoice
            invoice, _ = Invoice.objects.get_or_create(
                crew=cr,
                month=month_day
            )
            
            # Clear old invoice items if re-running
            invoice.invoiceitem_set.all().delete()
            
            # Calculate items and prepare for bulk create
            invoice_items = []
            total_for_crew = Decimal('0.00')
            
            for duty in duty_list:
                # Convert minutes -> hours
                hours = Decimal(duty.layover_time_minutes) / Decimal(60)
                if duty.arrival_airport and duty.arrival_airport.zone:
                    rate = duty.arrival_airport.zone.hourly_rate
                else:
                    rate = Decimal('0.00')
                
                line_amount = (hours * rate).quantize(Decimal('0.00'))
                total_for_crew += line_amount
                
                # Create invoice item
                item = InvoiceItem(
                    invoice=invoice,
                    duty=duty,
                    allowance_amount=line_amount
                )
                invoice_items.append(item)
            
            # Bulk create invoice items
            if invoice_items:
                InvoiceItem.objects.bulk_create(invoice_items)
            
            # Update invoice total
            invoice.total_amount = total_for_crew.quantize(Decimal('0.00'))
            invoice.save()

from django.db.models import Max
from django.db.models import Q

@login_required(login_url='login')
def crew_allowance_list(request):
    """
    Shows a paginated table of only those Invoices for the selected month,
    skipping any that are zero. If the user doesn't pick ?month=,
    we find the most recent month that has at least one invoice > 0.
    """
    month_str = request.GET.get('month')
    sort_order = request.GET.get('sort', 'asc') # default to ascending
    search_query = request.GET.get('search', '').strip() # search by first/last name
    # ------------------------
    # 1) If user picks a month, parse it
    # ------------------------
    if month_str:
        year, mo = map(int, month_str.split('-'))
        filter_month = date(year, mo, 1)
        
    else:
        # ------------------------
        # 2) Otherwise, find the "latest" month that has at least one > 0
        # ------------------------
        # We'll look at all distinct months in descending order:
        distinct_months = (
            Invoice.objects
            .values_list('month', flat=True)
            .distinct()
            .order_by('-month')
        )

        filter_month = None
        for m in distinct_months:
            # Check if this month has an Invoice with total_amount > 0
            has_nonzero = Invoice.objects.filter(month=m, total_amount__gt=0).exists()
            if has_nonzero:
                # This is our default month
                filter_month = m
                break

    # If we still have None => no data, or all zero
    if not filter_month:
        context = {
            'selected_month': None,
            'crew_data_list': [],
            'display_pages': [],
        }
        return render(request, 'aimsintegration/crew_allowance.html', context)

    # ------------------------
    # 3) Now fetch Invoices for that month, skipping total_amount=0
    # ------------------------
    invoices_qs = (
        Invoice.objects
        .filter(month=filter_month, total_amount__gt=0)  # skip zero
        .select_related('crew')
        .order_by('crew__crew_id')
    )

     # Apply search filter
    if search_query:
        invoices_qs = invoices_qs.filter(
            Q(crew__first_name__icontains=search_query) |
            Q(crew__last_name__icontains=search_query) |
            Q(crew__crew_id__icontains=search_query) |
            Q(crew__position__icontains=search_query)
        )

    # Apply sorting based on the sort order
    if sort_order == 'asc':
        invoices_qs = invoices_qs.order_by('total_amount')
    elif sort_order == 'desc':
        invoices_qs = invoices_qs.order_by('-total_amount')

    raw_crew_data_list = []
    for inv in invoices_qs:
        raw_crew_data_list.append({
            'crew': inv.crew,
            'total_amount': inv.total_amount,
            'invoice_id': inv.id,
            'invoice_month': inv.month,
        })

    # ------------------------
    # 4) Paginate
    # ------------------------
    paginator = Paginator(raw_crew_data_list, 10)
    page_number = request.GET.get('page', 1)
    try:
        crew_data_page = paginator.page(page_number)
    except PageNotAnInteger:
        crew_data_page = paginator.page(1)
    except EmptyPage:
        crew_data_page = paginator.page(paginator.num_pages)

    # ------------------------
    # 5) Windowed page range
    # ------------------------
    display_pages = get_display_pages(crew_data_page, num_links=2)

    context = {
        'selected_month': filter_month,
        'crew_data_list': crew_data_page,
        'display_pages': display_pages,
    }
    return render(request, 'aimsintegration/crew_allowance.html', context)





@login_required(login_url='login')
def compute_crew_allowance_for_month(crew, month_first_day):
    """
    On-the-fly calculation: sum each Duty's layover hours * arrival zone's rate.
    """
    year = month_first_day.year
    month = month_first_day.month

    duties = Duty.objects.filter(crew=crew, duty_date__year=year, duty_date__month=month)
    total = Decimal('0.00')
    for d in duties:
        if d.arrival_airport and d.arrival_airport.zone:
            rate = d.arrival_airport.zone.hourly_rate
        else:
            rate = Decimal('0.00')
        hours = Decimal(d.layover_time_minutes) / Decimal(60)
        total += hours * rate

    return total.quantize(Decimal('0.00'))

    
# Optimize crew_allowance_details function - use prefetch_related to reduce queries
@login_required(login_url='login')
def crew_allowance_details(request, crew_id, year, month):
    """
    Displays (or returns JSON) all duties for one crew in the chosen month.
    Optimized to reduce database queries with prefetch_related.
    """
    from decimal import Decimal

    cr = get_object_or_404(Crew, id=crew_id)
    filter_month = date(year, month, 1)

    try:
        # Use select_related and prefetch_related to reduce queries
        invoice = Invoice.objects.select_related('crew').get(crew=cr, month=filter_month)
        
        # Fetch all invoice items with related duties, departure and arrival airports in one query
        invoice_items = (invoice.invoiceitem_set
                         .select_related('duty__departure_airport', 
                                        'duty__arrival_airport', 
                                        'duty__arrival_airport__zone')
                         .filter(duty__layover_time_minutes__gt=0))
        
        duties_list = [item.duty for item in invoice_items]
        invoice_total = invoice.total_amount
    except Invoice.DoesNotExist:
        # Fallback: no invoice => fetch from Duty with optimized query
        duties_list = (Duty.objects
                       .select_related('departure_airport', 'arrival_airport', 'arrival_airport__zone')
                       .filter(crew=cr, duty_date__year=year, duty_date__month=month, 
                              layover_time_minutes__gt=0))
        
        invoice_total = compute_crew_allowance_for_month(cr, filter_month)

    # Build data structure more efficiently - avoid repeated calculations
    duties_data = []
    grand_total = Decimal('0.00')

    for d in duties_list:
        layover_hours = Decimal(d.layover_time_minutes) / Decimal(60)
        
        if d.arrival_airport and d.arrival_airport.zone:
            hourly_rate = d.arrival_airport.zone.hourly_rate
        else:
            hourly_rate = Decimal('0.00')

        line_amount = layover_hours * hourly_rate
        grand_total += line_amount

        duties_data.append({
            'duty_date': str(d.duty_date) if d.duty_date else "",
            'flight_number': d.flight_number,
            'departure': d.departure_airport.iata_code if d.departure_airport else "",
            'arrival': d.arrival_airport.iata_code if d.arrival_airport else "",
            'layover_hours': str(layover_hours.quantize(Decimal('0.00'))),
            'hourly_rate': str(hourly_rate.quantize(Decimal('0.00'))),
            'line_amount': str(line_amount.quantize(Decimal('0.00'))),
            'tail_number': d.tail_number,
        })

    # Use invoice_total for consistency
    final_total = invoice_total

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            'crew_info': f"{cr.crew_id} - {cr.first_name} {cr.last_name}",
            'duties': duties_data,
            'total_amount': str(final_total),
        }
        return JsonResponse(data)
    else:
        context = {
            'crew': cr,
            'filter_month': filter_month,
            'duties_data': duties_data,
            'final_total': final_total,
        }
        return render(request, 'aimsintegration/crew_allowance_details.html', context)

# somewhere in utils/pdf_utils.py
import io
from xhtml2pdf import pisa
# @login_required(login_url='login')
def convert_html_to_pdf(html_content):
    """
    Returns a bytes object of the PDF, or None if there's an error.
    """
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html_content.encode("utf-8")), result)
    if not pdf.err:
        return result.getvalue()
    return None



# Suppose you're in a view that handles "Generate Overall Payslip"


from django.db import connections
from decimal import Decimal
from datetime import date
from django.db.models import Max
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.templatetags.static import static
from .models import Invoice

from django.core.mail import EmailMessage
from django.contrib import messages
from django.conf import settings

@login_required(login_url='login')
def generate_overall_payslip(request):
    # 1) Read ?month=YYYY-MM
    month_str = request.GET.get('month')
    if not month_str:
        # Find the latest (max) month in your Invoice table if not provided
        last_invoice = (
            Invoice.objects
            .filter(total_amount__gt=0)
            .aggregate(max_month=Max('month'))
        )
        filter_month = last_invoice['max_month']
        if not filter_month:
            return HttpResponse("No invoices found at all in the database.", status=404)
    else:
        # Parse the user-provided month
        year, mo = map(int, month_str.split('-'))
        filter_month = date(year, mo, 1)

    # 2) Fetch Invoices for that month, skipping total_amount=0
    invoices = (
        Invoice.objects
        .filter(month=filter_month, total_amount__gt=0)
        .select_related('crew')
        .order_by('crew__crew_id')
    )

    if not invoices.exists():
        return HttpResponse("No nonzero invoices found for that month.", status=404)

    with connections['mssql'].cursor() as cursor:
        cursor.execute("""
            SELECT [Relational Exch_ Rate Amount]
            FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
            WHERE [Currency Code] = %s
            AND [Starting Date] = (
                SELECT MAX([Starting Date])
                FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [Currency Code] = %s
            );
        """, ['USD', 'USD'])
        row = cursor.fetchone()

    if not row:
        return HttpResponse("No exchange rate found from that query.", status=400)

    exchange_rate = Decimal(str(row[0]))

    # 4a) Gather all unique crew WB numbers and format them to WBXXXX
    crew_ids = list({inv.crew.crew_id for inv in invoices})
    wb_formatted_ids = [f"WB{int(cid):04d}" for cid in crew_ids]

    # 4b) Fetch bank details for these WB numbers
    employee_bank_data = {}
    if wb_formatted_ids:
        placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
        query = f"""
            SELECT [No_], [Bank Name], [Bank Account No]
            FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
            WHERE [No_] IN ({placeholders})
        """
        with connections['mssql'].cursor() as cursor:
            cursor.execute(query, wb_formatted_ids)
            rows = cursor.fetchall()

        for no_, bank_name, bank_account_no in rows:
            formatted_no = no_.strip()
            employee_bank_data[formatted_no] = {
                'bank_name': bank_name.strip() if bank_name else '',
                'bank_account_no': bank_account_no.strip() if bank_account_no else '',
            }

    # 4c) Build items list for the PDF table
    items = []
    for inv in invoices:
        wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
        bank_data = employee_bank_data.get(wb_formatted, {
            'bank_name': '',
            'bank_account_no': '',
        })

        usd_amount = inv.total_amount
        rwf_amount = (usd_amount * exchange_rate).quantize(Decimal('0.00'))
        bank_name = bank_data['bank_name'].strip() if bank_data['bank_name'] else '-'
        account_no = bank_data['bank_account_no'].strip() if bank_data['bank_account_no'] else '-'

        items.append({
            'wb_no': inv.crew.crew_id,
            'name': f"{inv.crew.first_name} {inv.crew.last_name}",
            'position': inv.crew.position,
            'usd_amount': usd_amount,
            'rwf_amount': f"{rwf_amount:,.2f}",  # Add commas to RWF
            'exchange_rate': exchange_rate,
            'bank_name': bank_name,
            'account_no': account_no,
        })

    # Calculate totals
    total_usd = sum([inv.total_amount for inv in invoices])
    total_rwf = sum([(inv.total_amount * exchange_rate).quantize(Decimal('0.00')) for inv in invoices])
    # Pass current date
    current_date = datetime.now().strftime("%B %d, %Y")  # Example: "January 19, 2025"
    # 5) Render to HTML via template
    context = {
        'items': items,
        'filter_month': filter_month,
        'exchange_rate': exchange_rate,
        'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
        'total_usd': f"{total_usd:,.2f}",  # Format with commas
        'total_rwf': f"{total_rwf:,.2f}",  # Format with commas
        'current_date': current_date,
    }
    html_string = render_to_string('aimsintegration/payslip_template.html', context)
        
    # 6) Convert the HTML to PDF
    pdf_file = convert_html_to_pdf(html_string)
    if not pdf_file:
        return HttpResponse("Error generating PDF.", status=500)

    # 7) Email the PDF to specified recipients
    try:
        # Create filename
        filename = f"Crew_Allowance_Payslip_{filter_month.strftime('%Y-%m')}.pdf"
        
        # Create email
        email = EmailMessage(
            subject=f'Crew Allowance Payslip - {filter_month.strftime("%B %Y")}',
            body=f'Please find attached the Crew Allowance Payslip for {filter_month.strftime("%B %Y")}.',
            from_email=settings.EMAIL_HOST_USER,
            to=['elie.kayitare@rwandair.com', 'saif.zawahreh@rwandair.com'],
        )
        
        # Attach PDF
        email.attach(filename, pdf_file, 'application/pdf')
        
        # Send email
        email.send(fail_silently=False)
        
        # Show success message to the user
        messages.success(request, f"Payslip successfully generated and emailed to recipients.")
    except Exception as e:
        # Log the error but continue to serve the PDF
        print(f"Error sending email: {str(e)}")
        messages.warning(request, f"Payslip generated but there was an error emailing: {str(e)}")

    # 8) Return PDF response to the user as well
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



# @login_required(login_url='login')
# def generate_usd_payslip(request):
#     # Get the month parameter
#     month_str = request.GET.get('month')
#     if not month_str:
#         last_invoice = Invoice.objects.filter(total_amount__gt=0, crew__position="CP").aggregate(max_month=Max('month'))
#         filter_month = last_invoice['max_month']
#         if not filter_month:
#             return HttpResponse("No invoices found for CP in the database.", status=404)
#     else:
#         year, mo = map(int, month_str.split('-'))
#         filter_month = date(year, mo, 1)

#     # Fetch invoices for CP (USD only)
#     cp_invoices = Invoice.objects.filter(
#         month=filter_month, total_amount__gt=0, crew__position="CP"
#     ).select_related('crew').order_by('crew__crew_id')

#     if not cp_invoices.exists():
#         return HttpResponse("No invoices found for CP.", status=404)

#     # Fetch bank details for these WB numbers
#     crew_ids = list({inv.crew.crew_id for inv in cp_invoices})
#     wb_formatted_ids = [f"WB{int(cid):04d}" for cid in crew_ids]

#     employee_bank_data = {}
#     if wb_formatted_ids:
#         placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
#         query = f"""
#             SELECT [No_], [Bank Name], [Bank Account No]
#             FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
#             WHERE [No_] IN ({placeholders})
#         """
#         with connections['mssql'].cursor() as cursor:
#             cursor.execute(query, wb_formatted_ids)
#             rows = cursor.fetchall()

#         for no_, bank_name, bank_account_no in rows:
#             formatted_no = no_.strip()
#             employee_bank_data[formatted_no] = {
#                 'bank_name': bank_name.strip() if bank_name else '-',
#                 'bank_account_no': bank_account_no.strip() if bank_account_no else '-',
#             }

#     # Prepare context
#     cp_items = []
#     for inv in cp_invoices:
#         wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
#         bank_data = employee_bank_data.get(wb_formatted, {
#             'bank_name': '-',
#             'bank_account_no': '-',
#         })

#         cp_items.append({
#             'wb_no': inv.crew.crew_id,
#             'name': f"{inv.crew.first_name} {inv.crew.last_name}",
#             'position': inv.crew.position,
#             'usd_amount': inv.total_amount,
#             'bank_name': bank_data['bank_name'],
#             'account_no': bank_data['bank_account_no'],
#         })

#     total_usd = sum([inv['usd_amount'] for inv in cp_items])

#     context = {
#         'items': cp_items,
#         'filter_month': filter_month,
#         'total_usd': f"{total_usd:,.2f}",
#         'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
#         'current_date': datetime.now().strftime("%B %d, %Y"),  # Example: "January 19, 2025"
#     }

#     # Render and generate PDF
#     html_string = render_to_string('aimsintegration/usd_payslip.html', context)
#     pdf_file = convert_html_to_pdf(html_string)
#     if not pdf_file:
#         return HttpResponse("Error generating CP PDF.", status=500)

#     # Return PDF response
#     response = HttpResponse(pdf_file, content_type='application/pdf')
#     filename = f"USD_Payslip_{filter_month.strftime('%Y-%m')}.pdf"
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#     return response





# @login_required(login_url='login')
# def generate_others_payslip(request):
#     # 1) Get the month parameter
#     month_str = request.GET.get('month')
#     if not month_str:
#         last_invoice = Invoice.objects.filter(total_amount__gt=0).exclude(crew__position="CP").aggregate(max_month=Max('month'))
#         filter_month = last_invoice['max_month']
#         if not filter_month:
#             return HttpResponse("No invoices found for others in the database.", status=404)
#     else:
#         year, mo = map(int, month_str.split('-'))
#         filter_month = date(year, mo, 1)

#     # 2) Fetch invoices for others
#     other_invoices = Invoice.objects.filter(
#         month=filter_month, total_amount__gt=0
#     ).exclude(crew__position="CP").select_related('crew').order_by('crew__crew_id')

#     if not other_invoices.exists():
#         return HttpResponse("No invoices found for others.", status=404)

#     # 3) Connect to MSSQL to get the exchange rate
#     # with connections['mssql'].cursor() as cursor:
#     #     cursor.execute("""
#     #         SELECT [Relational Exch_ Rate Amount]
#     #         FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
#     #         WHERE [Currency Code] = 'USD'
#     #           AND [Starting Date] = (
#     #               SELECT MAX([Starting Date])
#     #               FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
#     #               WHERE [Currency Code] = 'USD'
#     #           );
#     #     """)
#     #     row = cursor.fetchone()

#     with connections['mssql'].cursor() as cursor:
#         cursor.execute("""
#             SELECT [Relational Exch_ Rate Amount]
#             FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
#             WHERE [Currency Code] = %s
#             AND [Starting Date] = (
#                 SELECT MAX([Starting Date])
#                 FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
#                 WHERE [Currency Code] = %s
#             );
#         """, ['USD', 'USD'])
#         row = cursor.fetchone()

#     if not row:
#         return HttpResponse("No exchange rate found from that query.", status=400)

#     exchange_rate = Decimal(str(row[0]))

#     # 4a) Gather all unique crew WB numbers and format them
#     crew_ids = list({inv.crew.crew_id for inv in other_invoices})
#     wb_formatted_ids = [f"WB{int(cid):04d}" for cid in crew_ids]

#     # 4b) Fetch bank details for these WB numbers
#     employee_bank_data = {}
#     if wb_formatted_ids:
#         placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
#         query = f"""
#             SELECT [No_], [Bank Name], [Bank Account No]
#             FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
#             WHERE [No_] IN ({placeholders})
#         """
#         with connections['mssql'].cursor() as cursor:
#             cursor.execute(query, wb_formatted_ids)
#             rows = cursor.fetchall()

#         for no_, bank_name, bank_account_no in rows:
#             formatted_no = no_.strip()
#             employee_bank_data[formatted_no] = {
#                 'bank_name': bank_name.strip() if bank_name else '-',
#                 'bank_account_no': bank_account_no.strip() if bank_account_no else '-',
#             }

#     # 4c) Build items list for the PDF table
#     items = []
#     for inv in other_invoices:
#         wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
#         bank_data = employee_bank_data.get(wb_formatted, {
#             'bank_name': '-',
#             'bank_account_no': '-',
#         })

#         usd_amount = inv.total_amount
#         rwf_amount = (usd_amount * exchange_rate).quantize(Decimal('0.00'))
#         bank_name = bank_data['bank_name']
#         account_no = bank_data['bank_account_no']

#         items.append({
#             'wb_no': inv.crew.crew_id,
#             'name': f"{inv.crew.first_name} {inv.crew.last_name}",
#             'position': inv.crew.position,
#             'usd_amount': usd_amount,
#             'rwf_amount': f"{rwf_amount:,.2f}",
#             'exchange_rate': exchange_rate,
#             'bank_name': bank_name,
#             'account_no': account_no,
#         })

#     # Calculate totals
#     total_usd = sum([inv['usd_amount'] for inv in items])
#     total_rwf = sum([Decimal(inv['rwf_amount'].replace(',', '')) for inv in items])

#     # 5) Pass data to the context
#     current_date = datetime.now().strftime("%B %d, %Y")  # Example: "January 19, 2025"
#     context = {
#         'items': items,
#         'filter_month': filter_month,
#         'exchange_rate': exchange_rate,
#         'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
#         'total_usd': f"{total_usd:,.2f}",
#         'total_rwf': f"{total_rwf:,.2f}",
#         'current_date': current_date,
#     }

#     # 6) Render and generate the PDF
#     html_string = render_to_string('aimsintegration/others_payslip.html', context)
#     pdf_file = convert_html_to_pdf(html_string)
#     if not pdf_file:
#         return HttpResponse("Error generating Others PDF.", status=500)

#     # 7) Return the PDF response
#     response = HttpResponse(pdf_file, content_type='application/pdf')
#     filename = f"RWFs_Payslip_{filter_month.strftime('%Y-%m')}.pdf"
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#     return response


from django.core.mail import EmailMessage
from django.contrib import messages
from django.conf import settings
from decimal import Decimal
import io

# First, remove @login_required from this utility function
def convert_html_to_pdf(html_content):
    """
    Returns a bytes object of the PDF, or None if there's an error.
    """
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html_content.encode("utf-8")), result)
    if not pdf.err:
        return result.getvalue()
    return None

@login_required(login_url='login')
def generate_currency_payslips(request):
    """
    Single view that generates both the USD payslip (CP/FO) and RWF payslip (Others),
    emails both in a single email, and returns the requested one to the user.
    """
    # Determine which payslip the user wants to download (usd or rwf or both)
    payslip_type = request.GET.get('type', 'both') # Default is both
    
    # 1) Get the month parameter
    month_str = request.GET.get('month')
    if not month_str:
        # Find the latest month with invoices
        last_invoice = Invoice.objects.filter(total_amount__gt=0).aggregate(max_month=Max('month'))
        filter_month = last_invoice['max_month']
        if not filter_month:
            return HttpResponse("No invoices found in the database.", status=404)
    else:
        # Parse the user-provided month
        year, mo = map(int, month_str.split('-'))
        filter_month = date(year, mo, 1)

    # 2) Fetch ALL invoices for that month, skipping total_amount=0
    all_invoices = (
        Invoice.objects
        .filter(month=filter_month, total_amount__gt=0)
        .select_related('crew')
        .order_by('crew__position', 'crew__crew_id')
    )

    if not all_invoices.exists():
        return HttpResponse(f"No invoices found for {filter_month.strftime('%B %Y')}.", status=404)

    # 3) Split invoices into CP/FO and Others
    cp_fo_invoices = [inv for inv in all_invoices if inv.crew.position in ["CP", "FO"]]
    other_invoices = [inv for inv in all_invoices if inv.crew.position not in ["CP", "FO"]]

    # Check if we have invoices for both groups
    if payslip_type in ['usd', 'both'] and not cp_fo_invoices:
        if payslip_type == 'usd':
            return HttpResponse("No invoices found for CP or FO pilots.", status=404)
        else:
            messages.warning(request, "No CP/FO invoices found for this month, only generating Other Crew payslip.")
    
    if payslip_type in ['rwf', 'both'] and not other_invoices:
        if payslip_type == 'rwf':
            return HttpResponse("No invoices found for other crew positions.", status=404)
        else:
            messages.warning(request, "No Other Crew invoices found for this month, only generating CP/FO payslip.")

    # 4) Get exchange rate from MSSQL for the RWF payslip
    exchange_rate = None
    if other_invoices:
        with connections['mssql'].cursor() as cursor:
            cursor.execute("""
                SELECT [Relational Exch_ Rate Amount]
                FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [Currency Code] = %s
                AND [Starting Date] = (
                    SELECT MAX([Starting Date])
                    FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                    WHERE [Currency Code] = %s
                );
            """, ['USD', 'USD'])
            row = cursor.fetchone()

        if not row and payslip_type in ['rwf', 'both']:
            if payslip_type == 'rwf':
                return HttpResponse("No exchange rate found to generate RWF payslip.", status=400)
            else:
                messages.warning(request, "No exchange rate found, cannot generate RWF payslip.")
                other_invoices = []
        elif row:
            exchange_rate = Decimal(str(row[0]))

    # 5) Gather all unique crew WB numbers and format them
    all_crew_ids = list({inv.crew.crew_id for inv in all_invoices})
    wb_formatted_ids = [f"WB{int(cid):04d}" for cid in all_crew_ids]

    # 6) Fetch bank details for all crew members
    employee_bank_data = {}
    if wb_formatted_ids:
        placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
        query = f"""
            SELECT [No_], [Bank Name], [Bank Account No]
            FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
            WHERE [No_] IN ({placeholders})
        """
        with connections['mssql'].cursor() as cursor:
            cursor.execute(query, wb_formatted_ids)
            rows = cursor.fetchall()

        for no_, bank_name, bank_account_no in rows:
            formatted_no = no_.strip()
            employee_bank_data[formatted_no] = {
                'bank_name': bank_name.strip() if bank_name else '-',
                'bank_account_no': bank_account_no.strip() if bank_account_no else '-',
            }

    # 7) Create variables to hold the generated PDFs and filenames
    usd_pdf = None
    rwf_pdf = None
    usd_filename = f"USD_Payslip_CP_FO_{filter_month.strftime('%Y-%m')}.pdf"
    rwf_filename = f"RWF_Payslip_Others_{filter_month.strftime('%Y-%m')}.pdf"
    current_date = datetime.now().strftime("%B %d, %Y")
    
    # Track totals for the email
    cp_count = 0
    fo_count = 0
    total_usd_cp_fo = Decimal('0.00')
    position_summary = ""
    total_usd_others = Decimal('0.00')
    total_rwf_others = Decimal('0.00')

    # 8) Generate USD payslip for CP/FO if needed
    if cp_fo_invoices and payslip_type in ['usd', 'both']:
        # Build items list
        usd_items = []
        for inv in cp_fo_invoices:
            wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
            bank_data = employee_bank_data.get(wb_formatted, {
                'bank_name': '-',
                'bank_account_no': '-',
            })

            usd_items.append({
                'wb_no': inv.crew.crew_id,
                'name': f"{inv.crew.first_name} {inv.crew.last_name}",
                'position': inv.crew.position,
                'usd_amount': inv.total_amount,
                'bank_name': bank_data['bank_name'],
                'account_no': bank_data['bank_account_no'],
            })

        total_usd_cp_fo = sum([item['usd_amount'] for item in usd_items])
        
        # Get counts for email
        cp_count = sum(1 for item in usd_items if item['position'] == 'CP')
        fo_count = sum(1 for item in usd_items if item['position'] == 'FO')

        # Generate PDF
        usd_context = {
            'items': usd_items,
            'filter_month': filter_month,
            'total_usd': f"{total_usd_cp_fo:,.2f}",
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'current_date': current_date,
        }
        
        usd_html = render_to_string('aimsintegration/usd_payslip.html', usd_context)
        usd_pdf = convert_html_to_pdf(usd_html)
        if not usd_pdf and payslip_type == 'usd':
            return HttpResponse("Error generating CP/FO PDF.", status=500)

    # 9) Generate RWF payslip for Others if needed
    if other_invoices and exchange_rate and payslip_type in ['rwf', 'both']:
        # Build items list
        rwf_items = []
        for inv in other_invoices:
            wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
            bank_data = employee_bank_data.get(wb_formatted, {
                'bank_name': '-',
                'bank_account_no': '-',
            })

            usd_amount = inv.total_amount
            rwf_amount = (usd_amount * exchange_rate).quantize(Decimal('0.00'))

            rwf_items.append({
                'wb_no': inv.crew.crew_id,
                'name': f"{inv.crew.first_name} {inv.crew.last_name}",
                'position': inv.crew.position,
                'usd_amount': usd_amount,
                'rwf_amount': f"{rwf_amount:,.2f}",
                'exchange_rate': exchange_rate,
                'bank_name': bank_data['bank_name'],
                'account_no': bank_data['bank_account_no'],
            })

        total_usd_others = sum([item['usd_amount'] for item in rwf_items])
        total_rwf_others = sum([Decimal(item['rwf_amount'].replace(',', '')) for item in rwf_items])

        # Get position counts for email
        positions = sorted(set(item['position'] for item in rwf_items))
        position_counts = {pos: sum(1 for item in rwf_items if item['position'] == pos) for pos in positions}
        position_summary = ", ".join([f"{count} {pos}" for pos, count in position_counts.items()])

        # Generate PDF
        rwf_context = {
            'items': rwf_items,
            'filter_month': filter_month,
            'exchange_rate': exchange_rate,
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'total_usd': f"{total_usd_others:,.2f}",
            'total_rwf': f"{total_rwf_others:,.2f}",
            'current_date': current_date,
        }
        
        rwf_html = render_to_string('aimsintegration/others_payslip.html', rwf_context)
        rwf_pdf = convert_html_to_pdf(rwf_html)
        if not rwf_pdf and payslip_type == 'rwf':
            return HttpResponse("Error generating Others PDF.", status=500)

    # 10) Send a SINGLE email with BOTH PDFs (if generated)
    try:
        # Only send email if at least one PDF was generated
        if usd_pdf or rwf_pdf:
            month_year = filter_month.strftime("%B %Y")
            email_subject = f'RwandAir Crew Allowance Payslips - {month_year}'
            
            # Create detailed email body
            email_body = f"""
Dear Finance Team,

Please find attached the Crew Allowance Payslips for {month_year}. This email contains all payslips generated for this month.

"""
            # Add summary of attachments
            if usd_pdf and rwf_pdf:
                email_body += """ATTACHMENTS:
1. USD Payslip - For Captains and First Officers
2. RWF Payslip - For Other Crew Members

"""
            elif usd_pdf:
                email_body += """ATTACHMENTS:
1. USD Payslip - For Captains and First Officers

"""
            elif rwf_pdf:
                email_body += """ATTACHMENTS:
1. RWF Payslip - For Other Crew Members

"""

            # Add USD payslip details
            if usd_pdf:
                email_body += f"""
USD PAYSLIP SUMMARY:
- {cp_count} Captain(s)
- {fo_count} First Officer(s)
- Total USD Amount: {total_usd_cp_fo:,.2f}

"""

            # Add RWF payslip details
            if rwf_pdf:
                email_body += f"""
RWF PAYSLIP SUMMARY:
- Crew Positions: {position_summary}
- Total USD Amount: {total_usd_others:,.2f}
- Total RWF Amount: {total_rwf_others:,.2f}
- Exchange Rate: {exchange_rate}

"""
            
            # Add grand totals if both types exist
            if usd_pdf and rwf_pdf:
                grand_total_usd = total_usd_cp_fo + total_usd_others
                email_body += f"""
COMBINED TOTALS:
- Total USD Amount Across All Crew: {grand_total_usd:,.2f}
"""

            email_body += """
This is an automated message from the RwandAir Crew Allowance System.

Best Regards,
RwandAir Operations Team
"""
            
            # Create a single email with both attachments
            email = EmailMessage(
                subject=email_subject,
                body=email_body,
                from_email=settings.EMAIL_HOST_USER,
                to=['elie.kayitare@rwandair.com', 'saif.zawahreh@rwandair.com'],
            )
            
            # Attach both PDFs to the same email
            if usd_pdf:
                email.attach(usd_filename, usd_pdf, 'application/pdf')
            if rwf_pdf:
                email.attach(rwf_filename, rwf_pdf, 'application/pdf')
            
            # Send the single email with all attachments
            email.send(fail_silently=False)
            
            # Show success message with count of attachments
            if usd_pdf and rwf_pdf:
                messages.success(request, f"Both payslips (USD & RWF) successfully generated and emailed to the finance team.")
            elif usd_pdf:
                messages.success(request, f"USD payslip successfully generated and emailed to the finance team.")
            elif rwf_pdf:
                messages.success(request, f"RWF payslip successfully generated and emailed to the finance team.")
    except Exception as e:
        # Log the error but continue to serve the PDF
        print(f"Error sending email: {str(e)}")
        messages.warning(request, f"Payslips generated but there was an error emailing: {str(e)}")

    # 11) Return the requested PDF to the user
    if payslip_type == 'usd' and usd_pdf:
        response = HttpResponse(usd_pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{usd_filename}"'
        return response
    elif payslip_type == 'rwf' and rwf_pdf:
        response = HttpResponse(rwf_pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{rwf_filename}"'
        return response
    elif payslip_type == 'both':
        # If both requested, return the one that exists or the first one available
        if usd_pdf:
            response = HttpResponse(usd_pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{usd_filename}"'
            return response
        elif rwf_pdf:
            response = HttpResponse(rwf_pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{rwf_filename}"'
            return response
    
    # If we got here, we couldn't generate any PDFs
    return HttpResponse("No payslips could be generated.", status=404)


@login_required(login_url='login')
def generate_combined_payslips_email(request):
    """
    Generates both USD and RWF payslips in a single operation,
    sends a single email with both attachments, and returns either JSON (for AJAX)
    or redirects with a success message (for regular requests).
    """
    # Get the month parameter
    month_str = request.GET.get('month')
    if not month_str:
        # Find the latest month with invoices
        last_invoice = Invoice.objects.filter(total_amount__gt=0).aggregate(max_month=Max('month'))
        filter_month = last_invoice['max_month']
        if not filter_month:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'No invoices found in the database.'}, status=404)
            return HttpResponse("No invoices found in the database.", status=404)
    else:
        # Parse the user-provided month
        try:
            year, mo = map(int, month_str.split('-'))
            filter_month = date(year, mo, 1)
        except (ValueError, IndexError):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': f'Invalid month format: {month_str}'}, status=400)
            return HttpResponse(f"Invalid month format: {month_str}", status=400)

    # Fetch ALL invoices for that month, skipping total_amount=0
    all_invoices = (
        Invoice.objects
        .filter(month=filter_month, total_amount__gt=0)
        .select_related('crew')
        .order_by('crew__position', 'crew__crew_id')
    )

    if not all_invoices.exists():
        error_msg = f"No invoices found for {filter_month.strftime('%B %Y')}."
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=404)
        return HttpResponse(error_msg, status=404)

    # Split invoices into CP/FO and Others
    cp_fo_invoices = [inv for inv in all_invoices if inv.crew.position in ["CP", "FO"]]
    other_invoices = [inv for inv in all_invoices if inv.crew.position not in ["CP", "FO"]]

    # Check if we have invoices for both groups
    if not cp_fo_invoices:
        warning_msg = "No CP/FO invoices found for this month."
        if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
            messages.warning(request, warning_msg)
    
    if not other_invoices:
        warning_msg = "No other crew invoices found for this month."
        if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
            messages.warning(request, warning_msg)

    if not cp_fo_invoices and not other_invoices:
        error_msg = f"No valid invoices found for {filter_month.strftime('%B %Y')}."
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=404)
        return HttpResponse(error_msg, status=404)

    # Get exchange rate from MSSQL for the RWF payslip
    exchange_rate = None
    if other_invoices:
        with connections['mssql'].cursor() as cursor:
            cursor.execute("""
                SELECT [Relational Exch_ Rate Amount]
                FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [Currency Code] = %s
                AND [Starting Date] = (
                    SELECT MAX([Starting Date])
                    FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                    WHERE [Currency Code] = %s
                );
            """, ['USD', 'USD'])
            row = cursor.fetchone()

        if not row:
            warning_msg = "No exchange rate found, cannot generate RWF payslip."
            if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
                messages.warning(request, warning_msg)
            other_invoices = []
        else:
            exchange_rate = Decimal(str(row[0]))

    # Gather all unique crew WB numbers and format them
    all_crew_ids = list({inv.crew.crew_id for inv in all_invoices})
    wb_formatted_ids = [f"WB{int(cid):04d}" for cid in all_crew_ids]

    # Fetch bank details for all crew members
    employee_bank_data = {}
    if wb_formatted_ids:
        placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
        query = f"""
            SELECT [No_], [Bank Name], [Bank Account No]
            FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
            WHERE [No_] IN ({placeholders})
        """
        with connections['mssql'].cursor() as cursor:
            cursor.execute(query, wb_formatted_ids)
            rows = cursor.fetchall()

        for no_, bank_name, bank_account_no in rows:
            formatted_no = no_.strip()
            employee_bank_data[formatted_no] = {
                'bank_name': bank_name.strip() if bank_name else '-',
                'bank_account_no': bank_account_no.strip() if bank_account_no else '-',
            }

    # Create variables to hold the generated PDFs and filenames
    usd_pdf = None
    rwf_pdf = None
    usd_filename = f"USD_Payslip_CP_FO_{filter_month.strftime('%Y-%m')}.pdf"
    rwf_filename = f"RWF_Payslip_Others_{filter_month.strftime('%Y-%m')}.pdf"
    current_date = datetime.now().strftime("%B %d, %Y")
    
    # Track totals for the email and response
    cp_count = 0
    fo_count = 0
    total_usd_cp_fo = Decimal('0.00')
    position_summary = ""
    total_usd_others = Decimal('0.00')
    total_rwf_others = Decimal('0.00')

    # Generate USD payslip for CP/FO
    if cp_fo_invoices:
        # Build items list
        usd_items = []
        for inv in cp_fo_invoices:
            wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
            bank_data = employee_bank_data.get(wb_formatted, {
                'bank_name': '-',
                'bank_account_no': '-',
            })

            usd_items.append({
                'wb_no': inv.crew.crew_id,
                'name': f"{inv.crew.first_name} {inv.crew.last_name}",
                'position': inv.crew.position,
                'usd_amount': inv.total_amount,
                'bank_name': bank_data['bank_name'],
                'account_no': bank_data['bank_account_no'],
            })

        total_usd_cp_fo = sum([item['usd_amount'] for item in usd_items])
        
        # Get counts for email
        cp_count = sum(1 for item in usd_items if item['position'] == 'CP')
        fo_count = sum(1 for item in usd_items if item['position'] == 'FO')

        # Generate PDF
        usd_context = {
            'items': usd_items,
            'filter_month': filter_month,
            'total_usd': f"{total_usd_cp_fo:,.2f}",
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'current_date': current_date,
        }
        
        usd_html = render_to_string('aimsintegration/usd_payslip.html', usd_context)
        usd_pdf = convert_html_to_pdf(usd_html)
        if not usd_pdf:
            error_msg = "Error generating USD Payslip for CP/FO."
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)

    # Generate RWF payslip for Others
    if other_invoices and exchange_rate:
        # Build items list
        rwf_items = []
        for inv in other_invoices:
            wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
            bank_data = employee_bank_data.get(wb_formatted, {
                'bank_name': '-',
                'bank_account_no': '-',
            })

            usd_amount = inv.total_amount
            rwf_amount = (usd_amount * exchange_rate).quantize(Decimal('0.00'))

            rwf_items.append({
                'wb_no': inv.crew.crew_id,
                'name': f"{inv.crew.first_name} {inv.crew.last_name}",
                'position': inv.crew.position,
                'usd_amount': usd_amount,
                'rwf_amount': f"{rwf_amount:,.2f}",
                'exchange_rate': exchange_rate,
                'bank_name': bank_data['bank_name'],
                'account_no': bank_data['bank_account_no'],
            })

        total_usd_others = sum([item['usd_amount'] for item in rwf_items])
        total_rwf_others = sum([Decimal(item['rwf_amount'].replace(',', '')) for item in rwf_items])

        # Get position counts for email
        positions = sorted(set(item['position'] for item in rwf_items))
        position_counts = {pos: sum(1 for item in rwf_items if item['position'] == pos) for pos in positions}
        position_summary = ", ".join([f"{position_counts[pos]} {pos}" for pos in positions])

        # Generate PDF
        rwf_context = {
            'items': rwf_items,
            'filter_month': filter_month,
            'exchange_rate': exchange_rate,
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'total_usd': f"{total_usd_others:,.2f}",
            'total_rwf': f"{total_rwf_others:,.2f}",
            'current_date': current_date,
        }
        
        rwf_html = render_to_string('aimsintegration/others_payslip.html', rwf_context)
        rwf_pdf = convert_html_to_pdf(rwf_html)
        if not rwf_pdf:
            error_msg = "Error generating RWF Payslip for other crew positions."
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)

    # Send a SINGLE email with BOTH PDFs (if any were generated)
    email_status = False
    if usd_pdf or rwf_pdf:
        try:
            month_year = filter_month.strftime("%B %Y")
            email_subject = f'RwandAir Crew Allowance Payslips - {month_year}'
            
            # Create detailed email body
            email_body = f"""
Dear Finance Team,

Please find attached the Crew Allowance Payslips for {month_year}. This email contains all payslips generated for this month.

"""
            # Add summary of attachments
            if usd_pdf and rwf_pdf:
                email_body += """ATTACHMENTS:
1. USD Payslip - For Captains and First Officers
2. RWF Payslip - For Other Crew Members

"""
            elif usd_pdf:
                email_body += """ATTACHMENTS:
1. USD Payslip - For Captains and First Officers

"""
            elif rwf_pdf:
                email_body += """ATTACHMENTS:
1. RWF Payslip - For Other Crew Members

"""

            # Add USD payslip details
            if usd_pdf:
                email_body += f"""
USD PAYSLIP SUMMARY:
- {cp_count} Captain(s)
- {fo_count} First Officer(s)
- Total USD Amount: {total_usd_cp_fo:,.2f}

"""

            # Add RWF payslip details
            if rwf_pdf:
                email_body += f"""
RWF PAYSLIP SUMMARY:
- Crew Positions: {position_summary}
- Total USD Amount: {total_usd_others:,.2f}
- Total RWF Amount: {total_rwf_others:,.2f}
- Exchange Rate: {exchange_rate}

"""
            
            # Add grand totals if both types exist
            if usd_pdf and rwf_pdf:
                grand_total_usd = total_usd_cp_fo + total_usd_others
                email_body += f"""
COMBINED TOTALS:
- Total USD Amount Across All Crew: {grand_total_usd:,.2f}
"""

            email_body += """
This is an automated message from the RwandAir Crew Allowance System.

Best Regards,
RwandAir Operations Team
"""
            
            # Create a single email with both attachments
            email = EmailMessage(
                subject=email_subject,
                body=email_body,
                from_email=settings.EMAIL_HOST_USER,
                to=['elie.kayitare@rwandair.com', 'saif.zawahreh@rwandair.com'],
            )
            
            # Attach both PDFs to the same email
            if usd_pdf:
                email.attach(usd_filename, usd_pdf, 'application/pdf')
            if rwf_pdf:
                email.attach(rwf_filename, rwf_pdf, 'application/pdf')
            
            # Send the single email with all attachments
            email.send(fail_silently=False)
            email_status = True
            
        except Exception as e:
            # Log the error
            print(f"Error sending email: {str(e)}")
            error_msg = f"Payslips generated but could not send email: {str(e)}"
            if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
                messages.warning(request, error_msg)
            email_status = False
    
    # Return the appropriate response based on the request type
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # For AJAX requests, return a JSON response
        return JsonResponse({
            'success': True,
            'month': filter_month.strftime("%B %Y"),
            'has_usd': usd_pdf is not None,
            'has_rwf': rwf_pdf is not None,
            'cp_count': cp_count,
            'fo_count': fo_count,
            'other_positions': position_summary,
            'email_sent': email_status
        })
    else:
        # For regular requests, add a success message and redirect
        if usd_pdf or rwf_pdf:
            success_msg = "Payslips generated successfully."
            if email_status:
                success_msg += " Email sent to finance team."
            messages.success(request, success_msg)
        
        # Redirect back to the crew allowance list
        return redirect('crew_allowance_list')



@login_required(login_url='login')
def currency_payslip_download(request):
    """
    Downloads the requested currency payslip (USD or RWF) without re-generating the email.
    This is called after the email has already been sent via generate_combined_payslips_email.
    """
    payslip_type = request.GET.get('type', 'usd')  # Default to USD
    
    # Get month parameter
    month_str = request.GET.get('month')
    if not month_str:
        # Find the latest month with invoices
        last_invoice = Invoice.objects.filter(total_amount__gt=0).aggregate(max_month=Max('month'))
        filter_month = last_invoice['max_month']
        if not filter_month:
            return HttpResponse("No invoices found in the database.", status=404)
    else:
        # Parse the user-provided month
        try:
            year, mo = map(int, month_str.split('-'))
            filter_month = date(year, mo, 1)
        except (ValueError, IndexError):
            return HttpResponse(f"Invalid month format: {month_str}", status=400)
    
    # Generate filename based on payslip type
    if payslip_type == 'usd':
        filename = f"USD_Payslip_CP_FO_{filter_month.strftime('%Y-%m')}.pdf"
    else:
        filename = f"RWF_Payslip_Others_{filter_month.strftime('%Y-%m')}.pdf"
    
    # Filter invoices based on payslip type
    if payslip_type == 'usd':
        # Get CP/FO invoices
        invoices = (
            Invoice.objects
            .filter(month=filter_month, total_amount__gt=0, crew__position__in=["CP", "FO"])
            .select_related('crew')
            .order_by('crew__position', 'crew__crew_id')
        )
        
        if not invoices.exists():
            return HttpResponse(f"No CP/FO invoices found for {filter_month.strftime('%B %Y')}.", status=404)
    else:
        # Get other invoices
        invoices = (
            Invoice.objects
            .filter(month=filter_month, total_amount__gt=0)
            .exclude(crew__position__in=["CP", "FO"])
            .select_related('crew')
            .order_by('crew__position', 'crew__crew_id')
        )
        
        if not invoices.exists():
            return HttpResponse(f"No other crew invoices found for {filter_month.strftime('%B %Y')}.", status=404)
        
        # Get exchange rate if needed for RWF
        with connections['mssql'].cursor() as cursor:
            cursor.execute("""
                SELECT [Relational Exch_ Rate Amount]
                FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [Currency Code] = %s
                AND [Starting Date] = (
                    SELECT MAX([Starting Date])
                    FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                    WHERE [Currency Code] = %s
                );
            """, ['USD', 'USD'])
            row = cursor.fetchone()
        
        if not row:
            return HttpResponse("No exchange rate found to generate RWF payslip.", status=400)
        
        exchange_rate = Decimal(str(row[0]))
    
    # Fetch bank details for all crew members
    crew_ids = list({inv.crew.crew_id for inv in invoices})
    wb_formatted_ids = [f"WB{int(cid):04d}" for cid in crew_ids]
    
    employee_bank_data = {}
    if wb_formatted_ids:
        placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
        query = f"""
            SELECT [No_], [Bank Name], [Bank Account No]
            FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
            WHERE [No_] IN ({placeholders})
        """
        with connections['mssql'].cursor() as cursor:
            cursor.execute(query, wb_formatted_ids)
            rows = cursor.fetchall()

        for no_, bank_name, bank_account_no in rows:
            formatted_no = no_.strip()
            employee_bank_data[formatted_no] = {
                'bank_name': bank_name.strip() if bank_name else '-',
                'bank_account_no': bank_account_no.strip() if bank_account_no else '-',
            }
    
    # Generate the appropriate PDF
    current_date = datetime.now().strftime("%B %d, %Y")
    
    if payslip_type == 'usd':
        # Build items list for USD
        usd_items = []
        for inv in invoices:
            wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
            bank_data = employee_bank_data.get(wb_formatted, {
                'bank_name': '-',
                'bank_account_no': '-',
            })

            usd_items.append({
                'wb_no': inv.crew.crew_id,
                'name': f"{inv.crew.first_name} {inv.crew.last_name}",
                'position': inv.crew.position,
                'usd_amount': inv.total_amount,
                'bank_name': bank_data['bank_name'],
                'account_no': bank_data['bank_account_no'],
            })

        total_usd = sum([item['usd_amount'] for item in usd_items])
        
        # Generate USD PDF
        usd_context = {
            'items': usd_items,
            'filter_month': filter_month,
            'total_usd': f"{total_usd:,.2f}",
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'current_date': current_date,
        }
        
        html_string = render_to_string('aimsintegration/usd_payslip.html', usd_context)
        pdf_file = convert_html_to_pdf(html_string)
        
    else:
        # Build items list for RWF
        rwf_items = []
        for inv in invoices:
            wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
            bank_data = employee_bank_data.get(wb_formatted, {
                'bank_name': '-',
                'bank_account_no': '-',
            })

            usd_amount = inv.total_amount
            rwf_amount = (usd_amount * exchange_rate).quantize(Decimal('0.00'))

            rwf_items.append({
                'wb_no': inv.crew.crew_id,
                'name': f"{inv.crew.first_name} {inv.crew.last_name}",
                'position': inv.crew.position,
                'usd_amount': usd_amount,
                'rwf_amount': f"{rwf_amount:,.2f}",
                'exchange_rate': exchange_rate,
                'bank_name': bank_data['bank_name'],
                'account_no': bank_data['bank_account_no'],
            })

        total_usd = sum([item['usd_amount'] for item in rwf_items])
        total_rwf = sum([Decimal(item['rwf_amount'].replace(',', '')) for item in rwf_items])
        
        # Generate RWF PDF
        rwf_context = {
            'items': rwf_items,
            'filter_month': filter_month,
            'exchange_rate': exchange_rate,
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'total_usd': f"{total_usd:,.2f}",
            'total_rwf': f"{total_rwf:,.2f}",
            'current_date': current_date,
        }
        
        html_string = render_to_string('aimsintegration/others_payslip.html', rwf_context)
        pdf_file = convert_html_to_pdf(html_string)
    
    if not pdf_file:
        return HttpResponse(f"Error generating {payslip_type.upper()} PDF.", status=500)
    
    # Return the PDF file for download
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



@login_required(login_url='login')
def generate_individual_payslip(request, crew_id, year, month):
    """
    Generate and return an individual payslip PDF for a specific crew member.
    """
    try:
        crew = get_object_or_404(Crew, id=crew_id)
        filter_month = date(int(year), int(month), 1)
        
        # Get the invoice for this crew member and month
        try:
            invoice = Invoice.objects.get(crew=crew, month=filter_month)
            if invoice.total_amount <= 0:
                return HttpResponse("No allowance found for this crew member in this month.", status=404)
        except Invoice.DoesNotExist:
            return HttpResponse("No invoice found for this crew member in this month.", status=404)
        
        # Get exchange rate from MSSQL
        exchange_rate = None
        with connections['mssql'].cursor() as cursor:
            cursor.execute("""
                SELECT [Relational Exch_ Rate Amount]
                FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [Currency Code] = %s
                AND [Starting Date] = (
                    SELECT MAX([Starting Date])
                    FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                    WHERE [Currency Code] = %s
                );
            """, ['USD', 'USD'])
            row = cursor.fetchone()
            if row:
                exchange_rate = Decimal(str(row[0]))
        
        # Get all duties for this invoice
        invoice_items = invoice.invoiceitem_set.select_related(
            'duty__departure_airport', 
            'duty__arrival_airport', 
            'duty__arrival_airport__zone'
        ).filter(duty__layover_time_minutes__gt=0).all()
        
        duties_list = [item.duty for item in invoice_items]
        
        if not duties_list:
            return HttpResponse("No duties with layover found for this crew member.", status=404)
        
        # Get bank details for this crew member
        wb_formatted = f"WB{int(crew.crew_id):04d}"
        bank_data = {'bank_name': '-', 'bank_account_no': '-'}
        
        with connections['mssql'].cursor() as cursor:
            cursor.execute("""
                SELECT [Bank Name], [Bank Account No]
                FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [No_] = %s
            """, [wb_formatted])
            row = cursor.fetchone()
            if row:
                bank_data = {
                    'bank_name': row[0].strip() if row[0] else '-',
                    'bank_account_no': row[1].strip() if row[1] else '-',
                }
        
        # Build duties data for the template
        duties_data = []
        total_amount_usd = Decimal('0.00')
        
        for duty in duties_list:
            layover_hours = Decimal(duty.layover_time_minutes) / Decimal(60)
            if duty.arrival_airport and duty.arrival_airport.zone:
                hourly_rate = duty.arrival_airport.zone.hourly_rate
            else:
                hourly_rate = Decimal('0.00')
            
            line_amount = layover_hours * hourly_rate
            total_amount_usd += line_amount
            
            duties_data.append({
                'duty_date': duty.duty_date,
                'flight_number': duty.flight_number or '--',
                'departure': duty.departure_airport.iata_code if duty.departure_airport else '--',
                'arrival': duty.arrival_airport.iata_code if duty.arrival_airport else '--',
                'layover_hours': f"{layover_hours:.2f}",
                'hourly_rate': f"{hourly_rate:.2f}",
                'line_amount': f"{line_amount:.2f}",
                'tail_number': duty.tail_number or 'N/A',
            })
        
        # Calculate RWF amount
        total_amount_rwf = None
        if exchange_rate:
            total_amount_rwf = (total_amount_usd * exchange_rate).quantize(Decimal('0.00'))
        
        # Determine payment currency based on position
        payment_currency = 'USD' if crew.position in ['CP', 'FO'] else 'RWF'
        
        # Prepare context for individual payslip template
        context = {
            'crew': crew,
            'filter_month': filter_month,
            'duties_data': duties_data,
            'total_amount_usd': f"{total_amount_usd:.2f}",
            'total_amount_rwf': f"{total_amount_rwf:,.2f}" if total_amount_rwf else "N/A",
            'exchange_rate': exchange_rate,
            'payment_currency': payment_currency,
            'bank_name': bank_data['bank_name'],
            'account_no': bank_data['bank_account_no'],
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'current_date': datetime.now().strftime("%B %d, %Y"),
        }
        
        # Render the individual payslip template
        html_string = render_to_string('aimsintegration/individual_payslip.html', context)
        pdf_file = convert_html_to_pdf(html_string)
        
        if not pdf_file:
            return HttpResponse("Error generating individual payslip PDF.", status=500)
        
        # Return the PDF
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f"Individual_Payslip_{crew.crew_id}_{filter_month.strftime('%Y-%m')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating payslip: {str(e)}", status=500)


@login_required(login_url='login')
def email_individual_payslip(request, crew_id, year, month):
    """
    Generate and email an individual payslip to the crew member.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'}, status=405)
    
    try:
        crew = get_object_or_404(Crew, id=crew_id)
        filter_month = date(int(year), int(month), 1)
        
        # Get the invoice for this crew member and month
        try:
            invoice = Invoice.objects.get(crew=crew, month=filter_month)
            if invoice.total_amount <= 0:
                return JsonResponse({'success': False, 'error': 'No allowance found for this crew member'})
        except Invoice.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'No invoice found for this crew member'})
        
        # Get exchange rate from MSSQL
        exchange_rate = None
        with connections['mssql'].cursor() as cursor:
            cursor.execute("""
                SELECT [Relational Exch_ Rate Amount]
                FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [Currency Code] = %s
                AND [Starting Date] = (
                    SELECT MAX([Starting Date])
                    FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                    WHERE [Currency Code] = %s
                );
            """, ['USD', 'USD'])
            row = cursor.fetchone()
            if row:
                exchange_rate = Decimal(str(row[0]))
        
        # Get crew email from MSSQL employee table
        wb_formatted = f"WB{int(crew.crew_id):04d}"
        crew_email = None
        bank_data = {'bank_name': '-', 'bank_account_no': '-'}
        
        with connections['mssql'].cursor() as cursor:
            cursor.execute("""
                SELECT [Company E-Mail], [Bank Name], [Bank Account No]
                FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [No_] = %s
            """, [wb_formatted])
            row = cursor.fetchone()
            if row:
                crew_email = row[0].strip() if row[0] else None
                print("=========================================================")
                print(crew_email)
                print("=========================================================")
                bank_data = {
                    'bank_name': row[1].strip() if row[1] else '-',
                    'bank_account_no': row[2].strip() if row[2] else '-',
                }
        

        print("###########################################################")
        print(crew_email)
        print("###########################################################")
        if not crew_email:
            return JsonResponse({'success': False, 'error': 'No email address found for this crew member'})
        
        # Generate the same payslip data as in the print function
        invoice_items = invoice.invoiceitem_set.select_related(
            'duty__departure_airport', 
            'duty__arrival_airport', 
            'duty__arrival_airport__zone'
        ).filter(duty__layover_time_minutes__gt=0).all()
        
        duties_list = [item.duty for item in invoice_items]
        
        if not duties_list:
            return JsonResponse({'success': False, 'error': 'No duties with layover found'})
        
        # Build duties data for the template
        duties_data = []
        total_amount_usd = Decimal('0.00')
        
        for duty in duties_list:
            layover_hours = Decimal(duty.layover_time_minutes) / Decimal(60)
            if duty.arrival_airport and duty.arrival_airport.zone:
                hourly_rate = duty.arrival_airport.zone.hourly_rate
            else:
                hourly_rate = Decimal('0.00')
            
            line_amount = layover_hours * hourly_rate
            total_amount_usd += line_amount
            
            duties_data.append({
                'duty_date': duty.duty_date,
                'flight_number': duty.flight_number or '--',
                'departure': duty.departure_airport.iata_code if duty.departure_airport else '--',
                'arrival': duty.arrival_airport.iata_code if duty.arrival_airport else '--',
                'layover_hours': f"{layover_hours:.2f}",
                'hourly_rate': f"{hourly_rate:.2f}",
                'line_amount': f"{line_amount:.2f}",
                'tail_number': duty.tail_number or 'N/A',
            })
        
        # Calculate RWF amount
        total_amount_rwf = None
        if exchange_rate:
            total_amount_rwf = (total_amount_usd * exchange_rate).quantize(Decimal('0.00'))
        
        # Determine payment currency based on position
        payment_currency = 'USD' if crew.position in ['CP', 'FO'] else 'RWF'
        
        # Prepare context for individual payslip template
        context = {
            'crew': crew,
            'filter_month': filter_month,
            'duties_data': duties_data,
            'total_amount_usd': f"{total_amount_usd:.2f}",
            'total_amount_rwf': f"{total_amount_rwf:,.2f}" if total_amount_rwf else "N/A",
            'exchange_rate': exchange_rate,
            'payment_currency': payment_currency,
            'bank_name': bank_data['bank_name'],
            'account_no': bank_data['bank_account_no'],
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'current_date': datetime.now().strftime("%B %d, %Y"),
        }
        
        # Render and generate PDF
        html_string = render_to_string('aimsintegration/individual_payslip.html', context)
        pdf_file = convert_html_to_pdf(html_string)
        
        if not pdf_file:
            return JsonResponse({'success': False, 'error': 'Error generating PDF'})
        
        # Send email with currency information
        month_year = filter_month.strftime("%B %Y")
        email_subject = f'Your Crew Allowance Payslip - {month_year}'
        
        currency_info = f"${total_amount_usd:.2f}"
        if total_amount_rwf and payment_currency == 'RWF':
            currency_info += f" (RWF {total_amount_rwf:,.2f})"
        
        email_body = f"""
Dear {crew.first_name} {crew.last_name},

Please find attached your Crew Allowance Payslip for {month_year}.

Details:
- Total Amount: {currency_info}
- Payment Currency: {payment_currency}
- Period: {month_year}
- Generated: {datetime.now().strftime("%B %d, %Y")}

If you have any questions about your allowance, please contact the Operations team.

Best Regards,
RwandAir Operations Team
"""
        
        email = EmailMessage(
            subject=email_subject,
            body=email_body,
            from_email=settings.EMAIL_HOST_USER,
            to=[crew_email],
            cc=['elie.kayitare@rwandair.com', 'saif.zawahreh@rwandair.com'],
        )
        
        filename = f"Individual_Payslip_{crew.crew_id}_{filter_month.strftime('%Y-%m')}.pdf"
        email.attach(filename, pdf_file, 'application/pdf')
        
        # email.send(fail_silently=False)
        
        return JsonResponse({
            'success': True, 
            'crew_name': f"{crew.first_name} {crew.last_name}",
            'email': crew_email
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required(login_url='login')
def layover_setup(request):
    zones = Zone.objects.all().prefetch_related('airports')
    # Convert each zone’s airports into a list or handle in template
    return render(request, 'aimsintegration/layover_setup.html', {
        'zones': zones
    })




import zipfile
import io
from django.core.mail import EmailMessage
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from datetime import datetime
from decimal import Decimal
from django.db.models import Max

@login_required(login_url='login')
def generate_all_bank_payslips_zip_email(request):
    """
    Generate payslips for all banks, create ZIP file, and send via email.
    Also returns the ZIP file for download.
    """
    try:
        # 1) Get month parameter
        month_str = request.GET.get('month')
        
        if not month_str or month_str.lower() == 'null':
            last_invoice = (
                Invoice.objects
                .filter(total_amount__gt=0)
                .aggregate(max_month=Max('month'))
            )
            filter_month = last_invoice['max_month']
            if not filter_month:
                return JsonResponse({'success': False, 'error': 'No invoices found in database'})
        else:
            try:
                year, mo = map(int, month_str.split('-'))
                filter_month = date(year, mo, 1)
            except ValueError:
                return JsonResponse({'success': False, 'error': f'Invalid month format: {month_str}'})

        # 2) Get all bank names
        bank_names = get_all_bank_names_for_month(filter_month)
        
        if not bank_names:
            return JsonResponse({'success': False, 'error': 'No banks found for the selected month'})

        # 3) Get exchange rate
        exchange_rate = get_exchange_rate()
        if not exchange_rate:
            return JsonResponse({'success': False, 'error': 'No exchange rate found'})

        # 4) Create ZIP file in memory
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            total_banks_processed = 0
            total_crew_members = 0
            
            for bank_name in bank_names:
                try:
                    # Generate PDF for this bank
                    pdf_content, bank_stats = generate_single_bank_pdf(bank_name, filter_month, exchange_rate, request)
                    
                    if pdf_content:
                        # Create filename for this bank
                        safe_bank_name = bank_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
                        filename = f"{safe_bank_name}_Crew_Allowance_{filter_month.strftime('%Y-%m')}.pdf"
                        
                        # Add PDF to ZIP
                        zip_file.writestr(filename, pdf_content)
                        total_banks_processed += 1
                        total_crew_members += bank_stats['crew_count']
                        
                except Exception as e:
                    print(f"Error generating PDF for bank {bank_name}: {str(e)}")
                    continue

        zip_buffer.seek(0)
        zip_content = zip_buffer.getvalue()

        if total_banks_processed == 0:
            return JsonResponse({'success': False, 'error': 'No bank payslips could be generated'})

        # 5) Send email with ZIP attachment
        month_year = filter_month.strftime("%B %Y")
        email_subject = f'Bank-wise Crew Allowance Payslips - {month_year}'
        
        email_body = f"""
Dear Finance Team,

Please find attached the bank-wise crew allowance payslips for {month_year}.

Summary:
- Period: {month_year}
- Total Banks: {total_banks_processed}
- Total Crew Members: {total_crew_members}
- Generated: {datetime.now().strftime("%B %d, %Y at %H:%M")}

All payslips are organized by bank name in the attached ZIP file.

Best Regards,
RwandAir Crew Allowance System
"""

        # Create email
        email = EmailMessage(
            subject=email_subject,
            body=email_body,
            from_email=settings.EMAIL_HOST_USER,
            to=['elie.kayitare@rwandair.com', 'saif.zawahreh@rwandair.com'],
            cc=['elie.kayitare@rwandair.com'],  # Add other recipients as needed
        )
        
        # Attach ZIP file
        zip_filename = f"Bank_Payslips_{filter_month.strftime('%Y-%m')}.zip"
        email.attach(zip_filename, zip_content, 'application/zip')
        
        # Send email
        email.send(fail_silently=False)

        # 6) Return response
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # AJAX request - return JSON
            return JsonResponse({
                'success': True,
                'message': f'Bank payslips generated and emailed successfully!',
                'banks_processed': total_banks_processed,
                'crew_members': total_crew_members,
                'month': month_year
            })
        else:
            # Regular request - return ZIP file for download
            response = HttpResponse(zip_content, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            return response

    except Exception as e:
        error_msg = f'Error generating bank payslips: {str(e)}'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg})
        else:
            return HttpResponse(error_msg, status=500)


def get_all_bank_names_for_month(filter_month):
    """
    Get all unique bank names for crew members who have invoices in the specified month.
    """
    # Get all invoices for the month
    invoices = Invoice.objects.filter(
        month=filter_month, 
        total_amount__gt=0
    ).select_related('crew')
    
    if not invoices.exists():
        return []
    
    # Get unique crew IDs
    crew_ids = {inv.crew.crew_id for inv in invoices}
    wb_formatted_ids = [f"WB{int(cid):04d}" for cid in crew_ids]
    
    if not wb_formatted_ids:
        return []
    
    # Query MSSQL for bank names
    placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
    query = f"""
        SELECT DISTINCT [Bank Name]
        FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
        WHERE [No_] IN ({placeholders})
        AND [Bank Name] IS NOT NULL
        AND [Bank Name] != ''
    """
    
    with connections['mssql'].cursor() as cursor:
        cursor.execute(query, wb_formatted_ids)
        rows = cursor.fetchall()
        bank_names = [row[0].strip() for row in rows if row[0] and row[0].strip()]
    
    return sorted(bank_names)  # Return sorted list


def get_exchange_rate():
    """
    Get the latest USD exchange rate.
    """
    with connections['mssql'].cursor() as cursor:
        cursor.execute("""
            SELECT [Relational Exch_ Rate Amount]
            FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
            WHERE [Currency Code] = %s
            AND [Starting Date] = (
                SELECT MAX([Starting Date])
                FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [Currency Code] = %s
            );
        """, ['USD', 'USD'])
        row = cursor.fetchone()
        
    return Decimal(str(row[0])) if row else None


def generate_single_bank_pdf(bank_name, filter_month, exchange_rate, request):
    """
    Generate PDF for a single bank. Returns (pdf_content, stats_dict).
    """
    try:
        # Get invoices for the month
        invoices = (
            Invoice.objects
            .filter(month=filter_month, total_amount__gt=0)
            .select_related('crew')
            .order_by('crew__crew_id')
        )

        if not invoices.exists():
            return None, {'crew_count': 0}

        # Get crew IDs and bank data
        crew_ids = list({inv.crew.crew_id for inv in invoices})
        wb_formatted_ids = [f"WB{int(cid):04d}" for cid in crew_ids]
        employee_bank_data = {}

        if wb_formatted_ids:
            placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
            query = f"""
                SELECT [No_], [Bank Name], [Bank Account No]
                FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [No_] IN ({placeholders})
            """
            with connections['mssql'].cursor() as cursor:
                cursor.execute(query, wb_formatted_ids)
                rows = cursor.fetchall()

            for no_, b_name, account_no in rows:
                formatted_no = no_.strip()
                employee_bank_data[formatted_no] = {
                    'bank_name': b_name.strip() if b_name else '',
                    'account_no': account_no.strip() if account_no else '',
                }

        # Filter invoices for this specific bank
        items = []
        for inv in invoices:
            wb_formatted = f"WB{int(inv.crew.crew_id):04d}"
            bank_data = employee_bank_data.get(wb_formatted, {'bank_name': '', 'account_no': '-'})
            
            if bank_data['bank_name'] != bank_name:
                continue

            usd_amount = inv.total_amount
            rwf_amount = (usd_amount * exchange_rate).quantize(Decimal('0.00'))

            items.append({
                'wb_no': inv.crew.crew_id,
                'name': f"{inv.crew.first_name} {inv.crew.last_name}",
                'position': inv.crew.position,
                'usd_amount': usd_amount,
                'rwf_amount': f"{rwf_amount:,.2f}",
                'bank_name': bank_data['bank_name'],
                'account_no': bank_data['account_no'],
            })

        if not items:
            return None, {'crew_count': 0}

        # Calculate totals
        total_usd = sum([Decimal(item['usd_amount']) for item in items])
        total_rwf = sum([Decimal(item['rwf_amount'].replace(',', '')) for item in items])

        # Prepare context
        context = {
            'items': items,
            'filter_month': filter_month,
            'exchange_rate': exchange_rate,
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'total_usd': f"{total_usd:,.2f}",
            'total_rwf': f"{total_rwf:,.2f}",
            'current_date': datetime.now().strftime("%B %d, %Y"),
            'bank_name': bank_name,  # Add bank name to context
        }

        # Render HTML and convert to PDF
        html_string = render_to_string('aimsintegration/payslip_template.html', context)
        pdf_content = convert_html_to_pdf(html_string)
        
        return pdf_content, {'crew_count': len(items)}

    except Exception as e:
        print(f"Error generating PDF for bank {bank_name}: {str(e)}")
        return None, {'crew_count': 0}



# =====================================================================

# Individual Payslip Generation and Emailing zipped payslips

# =====================================================================

import zipfile
import io
from django.core.mail import EmailMessage
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from datetime import datetime, date
from decimal import Decimal
from django.db.models import Max
from django.db import connections
from django.templatetags.static import static
import time

@login_required(login_url='login')
def generate_all_individual_payslips_zip_email(request):
    """
    Generate individual payslips for all crew members, email each one to the crew member,
    create ZIP file, and send ZIP to finance team.
    """
    try:
        # 1) Get month parameter
        month_str = request.GET.get('month')
        
        if not month_str or month_str.lower() == 'null':
            last_invoice = (
                Invoice.objects
                .filter(total_amount__gt=0)
                .aggregate(max_month=Max('month'))
            )
            filter_month = last_invoice['max_month']
            if not filter_month:
                return JsonResponse({'success': False, 'error': 'No invoices found in database'})
        else:
            try:
                year, mo = map(int, month_str.split('-'))
                filter_month = date(year, mo, 1)
            except ValueError:
                return JsonResponse({'success': False, 'error': f'Invalid month format: {month_str}'})

        # 2) Get all crew members with non-zero invoices for this month
        invoices = Invoice.objects.filter(
            month=filter_month, 
            total_amount__gt=0
        ).select_related('crew').order_by('crew__crew_id')

        if not invoices.exists():
            return JsonResponse({'success': False, 'error': 'No crew members found with allowances for this month'})

        # 3) Get exchange rate
        exchange_rate = get_exchange_rate()
        if not exchange_rate:
            return JsonResponse({'success': False, 'error': 'No exchange rate found'})

        # 4) Get all crew emails and bank data from MSSQL
        crew_emails_and_banks = get_all_crew_emails_and_banks([inv.crew for inv in invoices])

        # 5) Generate individual payslips and collect stats
        zip_buffer = io.BytesIO()
        successful_emails = []
        failed_emails = []
        total_crew_processed = 0
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for invoice in invoices:
                try:
                    crew = invoice.crew
                    crew_data = crew_emails_and_banks.get(crew.crew_id, {})
                    
                    # Generate individual payslip PDF
                    pdf_content, crew_stats = generate_single_individual_payslip_pdf(
                        crew, filter_month, exchange_rate, request, crew_data
                    )
                    
                    if pdf_content:
                        # Create filename for this crew member
                        filename = f"Individual_Payslip_{crew.crew_id}_{filter_month.strftime('%Y-%m')}.pdf"
                        
                        # Add PDF to ZIP
                        zip_file.writestr(filename, pdf_content)
                        total_crew_processed += 1
                        
                        # Email individual payslip to crew member
                        if crew_data.get('email'):
                            try:
                                email_success = email_individual_payslip_to_crew(
                                    crew, filter_month, pdf_content, crew_data, crew_stats
                                )
                                if email_success:
                                    successful_emails.append(f"{crew.first_name} {crew.last_name} ({crew_data['email']})")
                                else:
                                    failed_emails.append(f"{crew.first_name} {crew.last_name} (email failed)")
                            except Exception as e:
                                failed_emails.append(f"{crew.first_name} {crew.last_name} (email error: {str(e)})")
                        else:
                            failed_emails.append(f"{crew.first_name} {crew.last_name} (no email address)")
                        
                        # Small delay to avoid overwhelming email server
                        time.sleep(0.1)
                        
                except Exception as e:
                    print(f"Error generating payslip for crew {crew.crew_id}: {str(e)}")
                    failed_emails.append(f"{crew.first_name} {crew.last_name} (generation failed)")
                    continue

        zip_buffer.seek(0)
        zip_content = zip_buffer.getvalue()

        if total_crew_processed == 0:
            return JsonResponse({'success': False, 'error': 'No individual payslips could be generated'})

        # 6) Send ZIP file to finance team
        finance_email_success = email_zip_to_finance_team(zip_content, filter_month, {
            'total_crew': total_crew_processed,
            'successful_emails': len(successful_emails),
            'failed_emails': len(failed_emails)
        })

        # 7) Return response with detailed statistics
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Individual payslips generated and processed successfully!',
                'total_crew_processed': total_crew_processed,
                'successful_emails': len(successful_emails),
                'failed_emails': len(failed_emails),
                'finance_email_sent': finance_email_success,
                'month': filter_month.strftime("%B %Y"),
                'details': {
                    'successful_emails': successful_emails[:10],  # First 10 for display
                    'failed_emails': failed_emails[:10] if failed_emails else []
                }
            })
        else:
            # Regular request - return ZIP file for download
            response = HttpResponse(zip_content, content_type='application/zip')
            zip_filename = f"All_Individual_Payslips_{filter_month.strftime('%Y-%m')}.zip"
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            return response

    except Exception as e:
        error_msg = f'Error generating individual payslips: {str(e)}'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg})
        else:
            return HttpResponse(error_msg, status=500)


def get_all_crew_emails_and_banks(crew_list):
    """
    Get email addresses and bank details for all crew members from MSSQL.
    """
    crew_data = {}
    
    if not crew_list:
        return crew_data
    
    # Format crew IDs for MSSQL query
    wb_formatted_ids = [f"WB{int(crew.crew_id):04d}" for crew in crew_list]
    
    if wb_formatted_ids:
        placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
        query = f"""
            SELECT [No_], [Company E-Mail], [Bank Name], [Bank Account No]
            FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
            WHERE [No_] IN ({placeholders})
        """
        
        with connections['mssql'].cursor() as cursor:
            cursor.execute(query, wb_formatted_ids)
            rows = cursor.fetchall()
            
            for no_, email, bank_name, account_no in rows:
                # Extract crew_id from WB format (e.g., WB0123 -> 123)
                crew_id = no_.strip().replace('WB', '').lstrip('0') or '0'
                crew_data[crew_id] = {
                    'email': email.strip() if email else None,
                    'bank_name': bank_name.strip() if bank_name else '-',
                    'account_no': account_no.strip() if account_no else '-',
                }
    
    return crew_data


def generate_single_individual_payslip_pdf(crew, filter_month, exchange_rate, request, crew_data):
    """
    Generate PDF for a single crew member's individual payslip.
    Returns (pdf_content, stats_dict).
    """
    try:
        # Get the invoice for this crew member and month
        try:
            invoice = Invoice.objects.get(crew=crew, month=filter_month)
            if invoice.total_amount <= 0:
                return None, {'total_amount_usd': 0}
        except Invoice.DoesNotExist:
            return None, {'total_amount_usd': 0}

        # Get all duties for this invoice
        invoice_items = invoice.invoiceitem_set.select_related(
            'duty__departure_airport', 
            'duty__arrival_airport', 
            'duty__arrival_airport__zone'
        ).filter(duty__layover_time_minutes__gt=0).all()
        
        duties_list = [item.duty for item in invoice_items]
        
        if not duties_list:
            return None, {'total_amount_usd': 0}

        # Build duties data for the template
        duties_data = []
        total_amount_usd = Decimal('0.00')
        
        for duty in duties_list:
            layover_hours = Decimal(duty.layover_time_minutes) / Decimal(60)
            if duty.arrival_airport and duty.arrival_airport.zone:
                hourly_rate = duty.arrival_airport.zone.hourly_rate
            else:
                hourly_rate = Decimal('0.00')
            
            line_amount = layover_hours * hourly_rate
            total_amount_usd += line_amount
            
            duties_data.append({
                'duty_date': duty.duty_date,
                'flight_number': duty.flight_number or '--',
                'departure': duty.departure_airport.iata_code if duty.departure_airport else '--',
                'arrival': duty.arrival_airport.iata_code if duty.arrival_airport else '--',
                'layover_hours': f"{layover_hours:.2f}",
                'hourly_rate': f"{hourly_rate:.2f}",
                'line_amount': f"{line_amount:.2f}",
                'tail_number': duty.tail_number or 'N/A',
            })

        # Calculate RWF amount
        total_amount_rwf = None
        if exchange_rate:
            total_amount_rwf = (total_amount_usd * exchange_rate).quantize(Decimal('0.00'))

        # Determine payment currency based on position
        payment_currency = 'USD' if crew.position in ['CP', 'FO'] else 'RWF'

        # Prepare context for individual payslip template
        context = {
            'crew': crew,
            'filter_month': filter_month,
            'duties_data': duties_data,
            'total_amount_usd': f"{total_amount_usd:.2f}",
            'total_amount_rwf': f"{total_amount_rwf:,.2f}" if total_amount_rwf else "N/A",
            'exchange_rate': exchange_rate,
            'payment_currency': payment_currency,
            'bank_name': crew_data.get('bank_name', '-'),
            'account_no': crew_data.get('account_no', '-'),
            'logo_path': request.build_absolute_uri(static('images/rwandair-logo.png')),
            'current_date': datetime.now().strftime("%B %d, %Y"),
        }

        # Render the individual payslip template
        html_string = render_to_string('aimsintegration/individual_payslip.html', context)
        pdf_content = convert_html_to_pdf(html_string)
        
        return pdf_content, {
            'total_amount_usd': float(total_amount_usd),
            'total_amount_rwf': float(total_amount_rwf) if total_amount_rwf else 0,
            'payment_currency': payment_currency
        }

    except Exception as e:
        print(f"Error generating individual payslip for crew {crew.crew_id}: {str(e)}")
        return None, {'total_amount_usd': 0}


def email_individual_payslip_to_crew(crew, filter_month, pdf_content, crew_data, crew_stats):
    """
    Email individual payslip to the crew member.
    """
    try:
        month_year = filter_month.strftime("%B %Y")
        email_subject = f'Your Crew Allowance Payslip - {month_year}'
        
        # Determine currency display
        if crew_stats['payment_currency'] == 'USD':
            currency_info = f"${crew_stats['total_amount_usd']:.2f}"
        else:
            currency_info = f"RWF {crew_stats['total_amount_rwf']:,.2f}"
            if crew_stats['total_amount_usd'] > 0:
                currency_info += f" (${crew_stats['total_amount_usd']:.2f})"

        email_body = f"""
Dear {crew.first_name} {crew.last_name},

Please find attached your Crew Allowance Payslip for {month_year}.

Details:
- Total Amount: {currency_info}
- Payment Currency: {crew_stats['payment_currency']}
- Period: {month_year}
- Generated: {datetime.now().strftime("%B %d, %Y")}

If you have any questions about your allowance, please contact the Operations team.

Best Regards,
RwandAir Operations Team
"""

        email = EmailMessage(
            subject=email_subject,
            body=email_body,
            from_email=settings.EMAIL_HOST_USER,
            to=[crew_data['email']]
            # cc=['elie.kayitare@rwandair.com', 'saif.zawahreh@rwandair.com'],
        )
        
        filename = f"Individual_Payslip_{crew.crew_id}_{filter_month.strftime('%Y-%m')}.pdf"
        email.attach(filename, pdf_content, 'application/pdf')
        
        email.send(fail_silently=False)
        return True
        
    except Exception as e:
        print(f"Error emailing payslip to crew {crew.crew_id}: {str(e)}")
        return False


def email_zip_to_finance_team(zip_content, filter_month, stats):
    """
    Email ZIP file containing all individual payslips to the finance team.
    """
    try:
        month_year = filter_month.strftime("%B %Y")
        email_subject = f'All Individual Crew Allowance Payslips - {month_year}'
        
        email_body = f"""
Dear Finance Team,

Please find attached all individual crew allowance payslips for {month_year}.

Summary:
- Period: {month_year}
- Total Crew Members: {stats['total_crew']}
- Successfully Emailed: {stats['successful_emails']}
- Failed Emails: {stats['failed_emails']}
- Generated: {datetime.now().strftime("%B %d, %Y at %H:%M")}

All individual payslips have been emailed to the respective crew members (where email addresses were available).
This ZIP file contains all generated payslips for your records.

Best Regards,
RwandAir Crew Allowance System
"""

        # Create email to finance team
        email = EmailMessage(
            subject=email_subject,
            body=email_body,
            from_email=settings.EMAIL_HOST_USER,
            to=['elie.kayitare@rwandair.com'],  # Update with actual finance team emails
            cc=['elie.kayitare@rwandair.com', 'saif.zawahreh@rwandair.com'],
        )
        
        # Attach ZIP file
        zip_filename = f"All_Individual_Payslips_{filter_month.strftime('%Y-%m')}.zip"
        email.attach(zip_filename, zip_content, 'application/zip')
        
        # Send email
        email.send(fail_silently=False)
        return True
        
    except Exception as e:
        print(f"Error emailing ZIP to finance team: {str(e)}")
        return False


from django.views.decorators.http import require_POST
from django.http import JsonResponse
import json
from .models import Zone, Airport

@require_POST
@login_required(login_url='login')
def create_zone(request):
    try:
        data = json.loads(request.body.decode('utf-8'))
        zone_name = data.get('zone_name')
        hourly_rate = data.get('hourly_rate')
        airport_codes = data.get('airports', [])

        if not zone_name or hourly_rate is None:
            return JsonResponse({'error': 'Missing zone_name or hourly_rate'}, status=400)

        # Create the Zone
        zone = Zone.objects.create(name=zone_name, hourly_rate=hourly_rate)

        # Create or update Airports for this zone
        for code in airport_codes:
            airport, created = Airport.objects.get_or_create(iata_code=code)
            airport.zone = zone
            airport.save()

        return JsonResponse({
            'message': 'Zone created successfully!',
            'zone': {
                'id': zone.id,
                'name': zone.name,
                'hourly_rate': str(zone.hourly_rate)
            }
        }, status=201)

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.shortcuts import get_object_or_404
from .models import Zone, Airport


@login_required(login_url='login')
def update_zone(request, zone_id):
    """
    Update a zone's details including its name, hourly rate, and associated airports.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            # Fetch the zone
            zone = get_object_or_404(Zone, id=zone_id)

            # Update Zone name and hourly rate
            zone.name = data.get("zone_name", zone.name)
            zone.hourly_rate = data.get("hourly_rate", zone.hourly_rate)
            zone.save()

            # Handle airport updates
            if "airports" in data:
                updated_airports = data["airports"]  # List of {id, iata_code, action}

                for airport_data in updated_airports:
                    airport_id = airport_data.get("id")
                    iata_code = airport_data.get("iata_code").strip().upper()
                    action = airport_data.get("action")  # "update", "delete", "add"

                    if action == "update" and airport_id:
                        airport = get_object_or_404(Airport, id=airport_id, zone=zone)
                        airport.iata_code = iata_code
                        airport.save()

                    elif action == "delete" and airport_id:
                        Airport.objects.filter(id=airport_id, zone=zone).delete()

                    elif action == "add":
                        Airport.objects.create(iata_code=iata_code, zone=zone)

            return JsonResponse({"message": "Zone updated successfully!"}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data"}, status=400)

    return JsonResponse({"error": "Invalid request method"}, status=405)





from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Zone

@login_required(login_url='login')
def get_zone_airports(request, zone_id):
    """
    Fetch all airports belonging to a specific zone.
    """
    zone = get_object_or_404(Zone, id=zone_id)
    airports = list(zone.airports.values("id", "iata_code"))
    return JsonResponse({"airports": airports})



from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from .models import Airport


@login_required(login_url='login')
def update_airport(request, airport_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            new_code = data.get("iata_code", "").strip().upper()
            airport = get_object_or_404(Airport, id=airport_id)
            airport.iata_code = new_code
            airport.save()
            return JsonResponse({"message": "Airport updated successfully!"})
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data"}, status=400)
    return JsonResponse({"error": "Invalid request method"}, status=405)


@login_required(login_url='login')
def delete_airport(request, airport_id):
    if request.method == "POST":
        airport = get_object_or_404(Airport, id=airport_id)
        airport.delete()
        return JsonResponse({"message": "Airport deleted successfully!"})
    return JsonResponse({"error": "Invalid request method"}, status=405)


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Zone


@login_required(login_url='login')
def delete_zone(request, zone_id):
    if request.method == "POST":
        zone = get_object_or_404(Zone, id=zone_id)
        zone.delete()  # This will cascade and delete associated airports
        return JsonResponse({"message": "Zone and associated airports deleted successfully!"})
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
import json
from .models import Zone, Airport


@login_required(login_url='login')
def add_airport(request, zone_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            airport_code = data.get("airport_code", "").strip().upper()
            if not airport_code:
                return JsonResponse({"error": "Airport code is required."}, status=400)
            
            zone = get_object_or_404(Zone, id=zone_id)
            # Create and associate a new Airport with the given zone
            Airport.objects.create(iata_code=airport_code, zone=zone)
            return JsonResponse({"message": "Airport added successfully!"}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data"}, status=400)
    return JsonResponse({"error": "Invalid request method"}, status=405)





# Start of Quality control

# Add this simple function to your views.py

# import csv
# from django.http import HttpResponse
# from django.contrib.auth.decorators import login_required
# from datetime import date
# from django.db.models import Max
# from decimal import Decimal

# @login_required(login_url='login')
# def generate_simple_csv_export(request):
#     """
#     Generate a simple CSV file combining individual and general payslip data.
#     """
#     # Get month parameter
#     month_str = request.GET.get('month')
    
#     if not month_str:
#         # Find the latest month with invoices
#         last_invoice = Invoice.objects.filter(total_amount__gt=0).aggregate(max_month=Max('month'))
#         filter_month = last_invoice['max_month']
#         if not filter_month:
#             return HttpResponse("No invoices found", status=404)
#     else:
#         try:
#             year, mo = map(int, month_str.split('-'))
#             filter_month = date(year, mo, 1)
#         except ValueError:
#             return HttpResponse("Invalid month format", status=400)

#     # Get all invoices for the month
#     invoices = Invoice.objects.filter(
#         month=filter_month, 
#         total_amount__gt=0
#     ).select_related('crew').order_by('crew__crew_id')

#     if not invoices.exists():
#         return HttpResponse(f"No data found for {filter_month.strftime('%B %Y')}", status=404)

#     # Get exchange rate
#     exchange_rate = get_exchange_rate()
#     if not exchange_rate:
#         exchange_rate = Decimal('1300.00')  # Default fallback

#     # Get crew bank data
#     crew_ids = [inv.crew.crew_id for inv in invoices]
#     wb_formatted_ids = [f"WB{int(cid):04d}" for cid in crew_ids]
    
#     employee_bank_data = {}
#     if wb_formatted_ids:
#         with connections['mssql'].cursor() as cursor:
#             placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
#             query = f"""
#                 SELECT [No_], [Bank Name], [Bank Account No]
#                 FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
#                 WHERE [No_] IN ({placeholders})
#             """
#             cursor.execute(query, wb_formatted_ids)
#             rows = cursor.fetchall()
            
#             for no_, bank_name, account_no in rows:
#                 employee_bank_data[no_.strip()] = {
#                     'bank_name': bank_name.strip() if bank_name else '',
#                     'account_no': account_no.strip() if account_no else '',
#                 }

#     # Create CSV response
#     response = HttpResponse(content_type='text/csv')
#     filename = f"Crew_Allowance_Export_{filter_month.strftime('%Y-%m')}.csv"
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
#     writer = csv.writer(response)
    
#     # Write headers
#     writer.writerow([
#         'WB Number',
#         # 'Crew ID', 
#         'First Name',
#         'Last Name', 
#         'Position',
#         # 'Bank Name',
#         # 'Account Number',
#         'Allowance USD',
#         'Allowance RWF',
#         'Payment Currency'
#     ])
    
#     # Write data rows
#     total_usd = Decimal('0.00')
#     total_rwf = Decimal('0.00')
    
#     for invoice in invoices:
#         crew = invoice.crew
#         wb_formatted = f"WB{int(crew.crew_id):04d}"
#         bank_data = employee_bank_data.get(wb_formatted, {'bank_name': '', 'account_no': ''})
        
#         usd_amount = invoice.total_amount
#         rwf_amount = usd_amount * exchange_rate
#         payment_currency = 'USD' if crew.position in ['CP', 'FO'] else 'RWF'
        
#         total_usd += usd_amount
#         total_rwf += rwf_amount
        
#         writer.writerow([
#             wb_formatted,
#             # crew.crew_id,
#             crew.first_name,
#             crew.last_name,
#             crew.position or '',
#             # bank_data['bank_name'],
#             # bank_data['account_no'],
#             f"{usd_amount:.2f}",
#             f"{rwf_amount:.2f}",
#             payment_currency
#         ])
    
#     # Write totals
#     writer.writerow([])
#     writer.writerow(['TOTALS', '', '', '', '', '', '', f"{total_usd:.2f}", f"{total_rwf:.2f}", ''])
#     writer.writerow(['Exchange Rate:', f"1 USD = {exchange_rate} RWF", '', '', '', '', '', '', '', ''])
#     writer.writerow(['Period:', filter_month.strftime('%B %Y'), '', '', '', '', '', '', '', ''])
    
#     return response


# Add this function to your views.py

import csv
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from datetime import date
from django.db.models import Max
from decimal import Decimal
from django.db import connections

def get_exchange_rate():
    """
    Get the latest USD exchange rate.
    """
    with connections['mssql'].cursor() as cursor:
        cursor.execute("""
            SELECT [Relational Exch_ Rate Amount]
            FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
            WHERE [Currency Code] = %s
            AND [Starting Date] = (
                SELECT MAX([Starting Date])
                FROM [RwandAir].[dbo].[RwandAir$Currency Exchange Rate$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [Currency Code] = %s
            );
        """, ['USD', 'USD'])
        row = cursor.fetchone()
             
        return Decimal(str(row[0])) if row else None

@login_required(login_url='login')
def generate_simple_csv_export(request):
    """
    Generate a CSV file with one row per crew member and each EARNING flight in its own column.
    Only includes flights that actually earn money (have layover time and zones).
    Removes duplicates to ensure each unique flight appears only once.
    """
    # Get month parameter
    month_str = request.GET.get('month')
    
    if not month_str:
        # Find the latest month with invoices
        last_invoice = Invoice.objects.filter(total_amount__gt=0).aggregate(max_month=Max('month'))
        filter_month = last_invoice['max_month']
        if not filter_month:
            return HttpResponse("No invoices found", status=404)
    else:
        try:
            year, mo = map(int, month_str.split('-'))
            filter_month = date(year, mo, 1)
        except ValueError:
            return HttpResponse("Invalid month format", status=400)

    # Get exchange rate with debugging
    exchange_rate = get_exchange_rate()
    print(f"DEBUG: Exchange rate from database: {exchange_rate}")
    
    if not exchange_rate:
        exchange_rate = Decimal('1430.00')  # Updated fallback to current rate
        print(f"DEBUG: Using fallback exchange rate: {exchange_rate}")
    else:
        print(f"DEBUG: Using database exchange rate: {exchange_rate}")

    # Get all crews who have duties in this month
    year = filter_month.year
    month = filter_month.month
    
    # Get ONLY duties that earn money (have layover time and zones)
    # Fixed: Match DISTINCT ON expressions with ORDER BY expressions
    earning_duties = Duty.objects.filter(
        duty_date__year=year,
        duty_date__month=month,
        layover_time_minutes__gt=0,  # Must have layover time
        arrival_airport__zone__isnull=False  # Must have a zone (to earn money)
    ).select_related(
        'crew',
        'departure_airport',
        'arrival_airport',
        'arrival_airport__zone'
    ).order_by(
        'crew_id', 'duty_date', 'flight_number', 'departure_airport', 'arrival_airport'
    ).distinct(
        'crew_id', 'duty_date', 'flight_number', 'departure_airport', 'arrival_airport'
    )

    if not earning_duties.exists():
        return HttpResponse(f"No earning flight duties found for {filter_month.strftime('%B %Y')}", status=404)

    # Get crew bank data for all crews
    crew_ids = list(set([duty.crew.crew_id for duty in earning_duties]))
    wb_formatted_ids = [f"WB{int(cid):04d}" for cid in crew_ids]
    
    employee_bank_data = {}
    if wb_formatted_ids:
        with connections['mssql'].cursor() as cursor:
            placeholders = ", ".join(["%s"] * len(wb_formatted_ids))
            query = f"""
                SELECT [No_], [Bank Name], [Bank Account No]
                FROM [RwandAir].[dbo].[RwandAir$Employee$04be3167-71f9-46b8-93ec-2d5e5e08bf9b]
                WHERE [No_] IN ({placeholders})
            """
            cursor.execute(query, wb_formatted_ids)
            rows = cursor.fetchall()
            
            for no_, bank_name, account_no in rows:
                employee_bank_data[no_.strip()] = {
                    'bank_name': bank_name.strip() if bank_name else '',
                    'account_no': account_no.strip() if account_no else '',
                }

    # Group earning duties by crew and remove duplicates manually as well
    crew_duties = {}
    max_earning_flights = 0
    
    for duty in earning_duties:
        crew_id = duty.crew.crew_id
        if crew_id not in crew_duties:
            crew_duties[crew_id] = {
                'crew': duty.crew,
                'duties': [],
                'seen_flights': set()  # Track unique flights
            }
        
        # Create a unique identifier for this flight
        flight_key = (
            duty.duty_date,
            duty.flight_number,
            duty.departure_airport.iata_code if duty.departure_airport else None,
            duty.arrival_airport.iata_code if duty.arrival_airport else None
        )
        
        # Only add if we haven't seen this exact flight before
        if flight_key not in crew_duties[crew_id]['seen_flights']:
            crew_duties[crew_id]['duties'].append(duty)
            crew_duties[crew_id]['seen_flights'].add(flight_key)
            
            # Track maximum earning flights for any crew member
            if len(crew_duties[crew_id]['duties']) > max_earning_flights:
                max_earning_flights = len(crew_duties[crew_id]['duties'])

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    filename = f"Crew_Earning_Flights_{filter_month.strftime('%Y-%m')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write header information
    writer.writerow([f"RwandAir Crew Earning Flights - {filter_month.strftime('%B %Y')}"])
    writer.writerow([f"Exchange Rate: 1 USD = {exchange_rate} RWF"])
    writer.writerow([f"Generated: {date.today().strftime('%Y-%m-%d')}"])
    writer.writerow([f"Note: Only unique flights that earn allowance money are included"])
    writer.writerow([])  # Empty row
    
    # Build dynamic headers - crew info + earning flight columns + totals
    headers = [
        'WB Number',
        'First Name', 
        'Last Name',
        'Position',
        # 'Bank Name',
        # 'Account Number'
    ]
    
    # Add earning flight columns (each earning flight gets its own column)
    for i in range(1, max_earning_flights + 1):
        headers.extend([
            f'Flight {i} Date',
            f'Flight {i} No',
            f'Flight {i} Route', 
            f'Flight {i} Layover (Hrs)',
            f'Flight {i} Rate (USD)',
            f'Flight {i} Amount (USD)'
        ])
    
    # Add total columns
    headers.extend([
        'Total Earning Flights',
        'Total Allowance USD',
        'Total Allowance RWF', 
        'Payment Currency'
    ])
    
    writer.writerow(headers)
    
    # Track totals
    grand_total_usd = Decimal('0.00')
    total_crew_count = 0
    
    # Debug: Track calculations
    debug_crew_totals = []
    
    # Write crew rows
    for crew_id, crew_data in crew_duties.items():
        crew = crew_data['crew']
        duties = crew_data['duties']
        
        wb_formatted = f"WB{int(crew.crew_id):04d}"
        bank_data = employee_bank_data.get(wb_formatted, {'bank_name': '', 'account_no': ''})
        
        # Start building the row with crew info
        row = [
            wb_formatted,
            crew.first_name,
            crew.last_name,
            crew.position or '',
            # bank_data['bank_name'],
            # bank_data['account_no']
        ]
        
        # Calculate totals for this crew
        total_usd = Decimal('0.00')
        
        # Add earning flight data columns
        for i in range(max_earning_flights):
            if i < len(duties):
                duty = duties[i]
                
                # Calculate layover hours and amount (all these flights should earn money)
                layover_hours = Decimal(duty.layover_time_minutes) / Decimal(60)
                hourly_rate = duty.arrival_airport.zone.hourly_rate
                line_amount = layover_hours * hourly_rate
                total_usd += line_amount
                
                # Debug: Print first few calculations
                if len(debug_crew_totals) < 3:
                    print(f"DEBUG Crew {crew.crew_id}: {layover_hours:.2f}h × ${hourly_rate:.2f} = ${line_amount:.2f}")
                
                # Create route string
                route = f"{duty.departure_airport.iata_code if duty.departure_airport else '---'}-{duty.arrival_airport.iata_code if duty.arrival_airport else '---'}"
                
                # Add flight data to row
                row.extend([
                    duty.duty_date.strftime('%Y-%m-%d'),
                    duty.flight_number or 'SIM',
                    route,
                    f"{layover_hours:.2f}",
                    f"{hourly_rate:.2f}",
                    f"{line_amount:.2f}"
                ])
            else:
                # Empty columns for crews with fewer earning flights
                row.extend(['', '', '', '', '', ''])
        
        # Convert to RWF
        total_rwf = total_usd * exchange_rate
        payment_currency = 'USD' if crew.position in ['CP', 'FO'] else 'RWF'
        
        # Debug: Track this crew's totals
        debug_crew_totals.append({
            'crew_id': crew.crew_id,
            'total_usd': total_usd,
            'total_rwf': total_rwf
        })
        
        # Print first few crew totals for debugging
        if len(debug_crew_totals) <= 3:
            print(f"DEBUG Crew {crew.crew_id}: Total USD = ${total_usd:.2f}, Total RWF = {total_rwf:.2f}")
        
        # Add totals to row
        row.extend([
            len(duties),  # Total earning flights
            f"{total_usd:.2f}",
            f"{total_rwf:.2f}",
            payment_currency
        ])
        
        writer.writerow(row)
        
        grand_total_usd += total_usd
        total_crew_count += 1
    
    # Write summary
    writer.writerow([])  # Empty row
    writer.writerow(['SUMMARY'])
    writer.writerow(['Total Crew Members with Earnings:', total_crew_count])
    writer.writerow(['Maximum Earning Flights per Crew:', max_earning_flights])
    
    # Calculate final totals with explicit precision
    final_rwf_total = grand_total_usd * exchange_rate
    
    # Debug what we're about to write
    usd_value = f"{grand_total_usd:.2f}"
    rwf_value = f"{final_rwf_total:.2f}"
    print(f"DEBUG: About to write USD: '{usd_value}'")
    print(f"DEBUG: About to write RWF: '{rwf_value}'")
    
    writer.writerow(['Grand Total USD:', usd_value])
    writer.writerow(['Grand Total RWF:', rwf_value])
    
    # Debug final calculation
    print(f"DEBUG FINAL: {grand_total_usd:.2f} USD × {exchange_rate:.2f} = {final_rwf_total:.2f} RWF")
    
    return response
#End of Quality control





#HUBSPOT

from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import BasicAuthentication, SessionAuthentication
from django.utils import timezone
from .models import TableauData
from .serializers import TableauDataSerializer
from .pagination import FlexiblePageSizePagination


# class TableauDataListView(generics.ListAPIView):
#     serializer_class = TableauDataSerializer
#     pagination_class = FlexiblePageSizePagination
    
#     # Password-based authentication
#     authentication_classes = [BasicAuthentication, SessionAuthentication]
#     permission_classes = [IsAuthenticated]
    
#     def get_queryset(self):
#         """
#         Filter records from TODAY upward (today and future data)
#         """
#         today = timezone.now().date()
#         return TableauData.objects.filter(
#             operation_day__gte=today
#         ).order_by('operation_day', 'std')


# class TableauDataListView(generics.ListAPIView):
#     serializer_class = TableauDataSerializer
#     pagination_class = FlexiblePageSizePagination
    
#     # Password-based authentication
#     authentication_classes = [BasicAuthentication, SessionAuthentication]
#     permission_classes = [IsAuthenticated]
    
#     def get_queryset(self):
#         """
#         Filter records from TODAY upward (today and future data)
#         Return only disrupted flights:
#         - Cancelled flights (cancelled_deleted = True)
#         - Delayed flights (departure_delay_time > 0)
#         """
#         today = timezone.now().date()
#         return TableauData.objects.filter(
#             models.Q(cancelled_deleted=True) | 
#             models.Q(departure_delay_time__gt=0)
#         ).order_by('operation_day', 'std').filter(
#             operation_day__gte=today
#         )


class TableauDataListView(generics.ListAPIView):
    serializer_class = TableauDataSerializer
    pagination_class = FlexiblePageSizePagination
    
    # Password-based authentication
    authentication_classes = [BasicAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        Filter records from TODAY upward (today and future data)
        
        Query parameter:
        - filter: SQL-like filter conditions
        
        Examples:
        ?filter=cancelled_deleted=true
        ?filter=departure_delay_time>0
        ?filter=cancelled_deleted=true OR departure_delay_time>0
        """
        today = timezone.now().date()
        queryset = TableauData.objects.filter(
            operation_day__gte=today
        )
        
        # Get filter parameter
        filter_param = self.request.query_params.get('filter', '')
        
        if filter_param:
            try:
                # Parse and apply filter
                queryset = self.apply_filter(queryset, filter_param)
            except Exception as e:
                # If filter parsing fails, return empty queryset or log error
                pass
        
        return queryset.order_by('operation_day', 'std')
    
    def apply_filter(self, queryset, filter_string):
        """
        Parse and apply filter conditions
        """
        # Handle OR conditions
        if ' OR ' in filter_string.upper():
            conditions = filter_string.upper().split(' OR ')
            q_objects = models.Q()
            
            for condition in conditions:
                condition = condition.strip()
                q_obj = self.parse_condition(condition)
                if q_obj:
                    q_objects |= q_obj
            
            if q_objects:
                queryset = queryset.filter(q_objects)
        
        # Handle single condition
        else:
            q_obj = self.parse_condition(filter_string)
            if q_obj:
                queryset = queryset.filter(q_obj)
        
        return queryset
    
    def parse_condition(self, condition):
        """
        Parse individual filter condition
        """
        condition = condition.strip()
        
        # Handle cancelled_deleted=true
        if 'cancelled_deleted=true' in condition.lower():
            return models.Q(cancelled_deleted=True)
        
        # Handle cancelled_deleted=false
        elif 'cancelled_deleted=false' in condition.lower():
            return models.Q(cancelled_deleted=False)
        
        # Handle departure_delay_time>0
        elif 'departure_delay_time>0' in condition.lower():
            return models.Q(departure_delay_time__gt=0)
        
        # Handle departure_delay_time>X (specific number)
        elif 'departure_delay_time>' in condition.lower():
            try:
                value = condition.lower().split('departure_delay_time>')[1]
                delay_value = int(value)
                return models.Q(departure_delay_time__gt=delay_value)
            except (IndexError, ValueError):
                pass
        
        # Handle departure_delay_time=X
        elif 'departure_delay_time=' in condition.lower():
            try:
                value = condition.lower().split('departure_delay_time=')[1]
                delay_value = int(value)
                return models.Q(departure_delay_time=delay_value)
            except (IndexError, ValueError):
                pass
        
        return None
    


# ------------------------------------------------------------------------------------

# Qatar APIS

# ------------------------------------------------------------------------------------------

from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import Count, Q
from .models import QatarFlightDetails, QatarCrewDetail
import logging

logger = logging.getLogger(__name__)


def qatar_apis_dashboard(request):
    """
    Main dashboard view for Qatar APIS records.
    Shows all flight assignments with their status and crew counts.
    """
    try:
        # Get unique flight assignments (group by flight, date, direction, and filename)
        # We want one row per APIS file generated
        apis_records = QatarFlightDetails.objects.select_related('flight').filter(
            apis_filename__isnull=False  # Only show records that have been included in a file
        ).distinct('flight', 'dep_date_utc', 'direction', 'apis_filename').order_by(
            'flight', 'dep_date_utc', 'direction', 'apis_filename', '-status_updated_at'
        )
        
        # Calculate crew count for each record
        for record in apis_records:
            record.crew_count = QatarFlightDetails.objects.filter(
                flight=record.flight,
                dep_date_utc=record.dep_date_utc,
                apis_filename=record.apis_filename
            ).count()
        
        # Summary statistics
        total_flights = QatarFlightDetails.objects.filter(
            apis_filename__isnull=False
        ).values('flight', 'dep_date_utc').distinct().count()
        
        total_crew = QatarFlightDetails.objects.filter(
            apis_filename__isnull=False
        ).count()
        
        # Status counts
        status_counts = QatarFlightDetails.objects.filter(
            apis_filename__isnull=False
        ).values('status').annotate(count=Count('status'))
        
        status_dict = {item['status']: item['count'] for item in status_counts}
        
        context = {
            'apis_records': apis_records,
            'total_flights': total_flights,
            'total_crew': total_crew,
            'status_counts': status_dict,
        }
        
        return render(request, 'aimsintegration/qatar_apis_dashboard.html', context)
        
    except Exception as e:
        logger.error(f"Error in qatar_apis_dashboard: {e}")
        return render(request, 'aimsintegration/qatar_apis_dashboard.html', {
            'apis_records': [],
            'total_flights': 0,
            'total_crew': 0,
            'status_counts': {},
        })


def qatar_apis_details(request, record_id):
    """
    API endpoint to fetch detailed crew information for a specific flight assignment.
    Returns JSON with flight info and all crew members.
    """
    try:
        # Get the main record
        record = QatarFlightDetails.objects.select_related('flight').get(id=record_id)
        
        # Get all crew for this flight with the same APIS filename
        crew_assignments = QatarFlightDetails.objects.filter(
            flight=record.flight,
            dep_date_utc=record.dep_date_utc,
            apis_filename=record.apis_filename
        ).select_related('flight')
        
        # Collect crew details
        crew_members = []
        for crew_assignment in crew_assignments:
            crew_detail = QatarCrewDetail.objects.filter(crew_id=crew_assignment.crew_id).first()
            
            if crew_detail:
                crew_members.append({
                    'crew_id': crew_detail.crew_id,
                    'surname': crew_detail.surname or '',
                    'firstname': crew_detail.firstname or '',
                    'middlename': crew_detail.middlename or '',
                    'passport_number': crew_detail.passport_number or '',
                    'nationality_code': crew_detail.nationality_code or '',
                    'birth_date': crew_detail.birth_date.strftime('%Y-%m-%d') if crew_detail.birth_date else '--',
                    'sex': crew_detail.sex or '--',
                    'passport_expiry': crew_detail.passport_expiry.strftime('%Y-%m-%d') if crew_detail.passport_expiry else '--',
                })
        
        # Prepare response data
        data = {
            'flight_no': record.flight.flight_no,
            'tail_no': record.tail_no,
            'dep_date': record.dep_date_utc.strftime('%Y-%m-%d'),
            'route': f"{record.flight.dep_code_iata} → {record.flight.arr_code_iata}",
            'direction': record.direction,
            'status': record.get_status_display(),
            'apis_filename': record.apis_filename or '--',
            'std_utc': record.std_utc.strftime('%H:%M') if record.std_utc else '--',
            'sta_utc': record.sta_utc.strftime('%H:%M') if record.sta_utc else '--',
            'crew_members': crew_members,
        }
        
        return JsonResponse(data)
        
    except QatarFlightDetails.DoesNotExist:
        return JsonResponse({'error': 'Record not found'}, status=404)
    except Exception as e:
        logger.error(f"Error in qatar_apis_details: {e}")
        return JsonResponse({'error': 'Internal server error'}, status=500)
    



# ================================================================
# JEPPESSEN GENERAL DECLARATION (GD) - ENHANCED WITH FLITELINK
# ================================================================

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.db.models import Count, Q
from datetime import date, timedelta
from .models import (
    JEPPESSENGDFlight, 
    JEPPESSENGDCrew, 
    JEPPESSENGDProcessingLog, 
    JEPPESSENGDCrewDetail,
    FlitelinkAPILog
)
import logging

logger = logging.getLogger(__name__)


def jeppessen_dashboard(request):
    """
    Enhanced Jeppessen General Declaration dashboard with Flitelink integration.
    Shows GD flights with PIC/SIC tracking AND Flitelink submission status.
    """
    # Get date range
    days_back = int(request.GET.get('days', 30))
    start_date = date.today() - timedelta(days=days_back)
    
    # Get GD flights
    gd_flights = JEPPESSENGDFlight.objects.filter(
        flight_date__gte=start_date
    ).order_by('-flight_date', 'flight_no')
    
    # Today's statistics
    today = date.today()
    today_flights = JEPPESSENGDFlight.objects.filter(flight_date=today)
    today_crew = JEPPESSENGDCrew.objects.filter(flight_date=today)
    
    # PIC/SIC stats
    today_pic = today_crew.filter(is_pic=True).count()
    today_sic = today_crew.filter(is_sic=True).count()
    
    # Email stats
    today_crew_with_email = today_crew.exclude(email='').exclude(email__isnull=True).count()
    today_crew_without_email = today_crew.filter(Q(email='') | Q(email__isnull=True)).count()
    
    # Calculate email success rate
    email_success_rate = 0
    if today_crew.count() > 0:
        email_success_rate = (today_crew_with_email / today_crew.count()) * 100
    
    # Position breakdown
    position_counts = today_crew.values('position').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # ✨ NEW: Flitelink statistics
    flitelink_not_submitted = gd_flights.filter(flitelink_status='NOT_SUBMITTED').count()
    flitelink_pending = gd_flights.filter(flitelink_status='PENDING').count()
    flitelink_queued = gd_flights.filter(flitelink_status='QUEUED').count()
    flitelink_completed = gd_flights.filter(flitelink_status='COMPLETED').count()
    flitelink_failed = gd_flights.filter(flitelink_status='FAILED').count()
    
    # Flitelink success rate
    total_submitted = flitelink_queued + flitelink_completed + flitelink_failed + flitelink_pending
    flitelink_success_rate = 0
    if total_submitted > 0:
        flitelink_success_rate = (flitelink_completed / total_submitted) * 100
    
    # Overall stats
    total_flights = gd_flights.count()
    total_crew = JEPPESSENGDCrew.objects.filter(
        flight_date__gte=start_date
    ).count()
    
    # Prepare flight data with Flitelink status
    flight_data = []
    for gd_flight in gd_flights:
        crew_count = gd_flight.crew_assignments.count()
        pic = gd_flight.crew_assignments.filter(is_pic=True).first()
        sic = gd_flight.crew_assignments.filter(is_sic=True).first()
        
        # Get PIC/SIC names
        pic_name = None
        sic_name = None
        
        if pic:
            pic_detail = JEPPESSENGDCrewDetail.objects.filter(crew_id=pic.crew_id).first()
            if pic_detail:
                pic_name = pic_detail.full_name or f"{pic_detail.surname}, {pic_detail.firstname}"
        
        if sic:
            sic_detail = JEPPESSENGDCrewDetail.objects.filter(crew_id=sic.crew_id).first()
            if sic_detail:
                sic_name = sic_detail.full_name or f"{sic_detail.surname}, {sic_detail.firstname}"
        
        flight_data.append({
            'id': gd_flight.id,
            'flight_no': gd_flight.flight_no,
            'flight_date': gd_flight.flight_date,
            'origin_iata': gd_flight.origin_iata,
            'destination_iata': gd_flight.destination_iata,
            'tail_no': gd_flight.tail_no,
            'std_utc': gd_flight.std_utc,
            'sta_utc': gd_flight.sta_utc,
            'crew_count': crew_count,
            'pic_id': pic.crew_id if pic else None,
            'pic_name': pic_name,
            'sic_id': sic.crew_id if sic else None,
            'sic_name': sic_name,
            'processed_at': gd_flight.processed_at,
            # ✨ NEW: Flitelink data
            'flitelink_status': gd_flight.flitelink_status,
            'flitelink_status_display': gd_flight.get_flitelink_status_display(),
            'flitelink_request_id': str(gd_flight.flitelink_request_id) if gd_flight.flitelink_request_id else None,
            'flitelink_submitted_at': gd_flight.flitelink_submitted_at,
            'flitelink_completed_at': gd_flight.flitelink_completed_at,
            'flitelink_error_message': gd_flight.flitelink_error_message,
            'flitelink_retry_count': gd_flight.flitelink_retry_count,
            'can_submit_to_flitelink': gd_flight.can_submit_to_flitelink,
        })
    
    # Recent logs
    recent_logs = JEPPESSENGDProcessingLog.objects.all()[:5]
    
    # ✨ NEW: Recent Flitelink API logs
    recent_api_logs = FlitelinkAPILog.objects.select_related('gd_flight').all()[:10]
    
    context = {
        'flights': flight_data,
        'total_flights': total_flights,
        'total_crew': total_crew,
        'today_flights': today_flights.count(),
        'today_crew': today_crew.count(),
        'today_pic': today_pic,
        'today_sic': today_sic,
        'today_crew_with_email': today_crew_with_email,
        'today_crew_without_email': today_crew_without_email,
        'email_success_rate': email_success_rate,
        'position_counts': position_counts,
        'recent_logs': recent_logs,
        'days_back': days_back,
        'start_date': start_date,
        # ✨ NEW: Flitelink stats
        'flitelink_not_submitted': flitelink_not_submitted,
        'flitelink_pending': flitelink_pending,
        'flitelink_queued': flitelink_queued,
        'flitelink_completed': flitelink_completed,
        'flitelink_failed': flitelink_failed,
        'flitelink_success_rate': flitelink_success_rate,
        'recent_api_logs': recent_api_logs,
    }
    
    return render(request, 'aimsintegration/jeppessen_dashboard.html', context)



def jeppessen_flight_details(request, flight_id):
    """
    Enhanced API endpoint with Flitelink information.
    """
    try:
        gd_flight = JEPPESSENGDFlight.objects.get(id=flight_id)
        
        crew_assignments = gd_flight.crew_assignments.all().order_by(
            '-is_pic', '-is_sic', 'position'
        )
        
        crew_data = []
        for crew in crew_assignments:
            crew_detail = JEPPESSENGDCrewDetail.objects.filter(crew_id=crew.crew_id).first()
            
            crew_data.append({
                'crew_id': crew.crew_id,
                'position': crew.position,
                'position_display': crew.get_position_display(),
                'role': crew.role or '--',
                'is_pic': crew.is_pic,
                'is_sic': crew.is_sic,
                'name': crew_detail.full_name if crew_detail and crew_detail.full_name else '--',
                'email': crew.email or '--',
                'passport': crew_detail.passport_number if crew_detail else '--',
                'nationality': crew_detail.nationality_code if crew_detail else '--',
                'birth_date': crew_detail.birth_date.strftime('%Y-%m-%d') if crew_detail and crew_detail.birth_date else '--',
                'gender': crew_detail.sex if crew_detail else '--',
            })
        
        # ✨ NEW: Flitelink API logs for this flight
        api_logs = FlitelinkAPILog.objects.filter(gd_flight=gd_flight).order_by('-request_time')
        
        api_logs_data = []
        for log in api_logs:
            api_logs_data.append({
                'request_type': log.get_request_type_display(),
                'http_method': log.http_method,
                'response_status': log.response_status_code,
                'success': log.success,
                'duration_ms': log.duration_ms,
                'request_time': log.request_time.strftime('%Y-%m-%d %H:%M:%S'),
                'error_message': log.error_message,
            })
        
        data = {
            'flight_no': gd_flight.flight_no,
            'flight_date': gd_flight.flight_date.strftime('%Y-%m-%d'),
            'origin_iata': gd_flight.origin_iata,
            'origin_icao': gd_flight.origin_icao or '--',
            'destination_iata': gd_flight.destination_iata,
            'destination_icao': gd_flight.destination_icao or '--',
            'tail_no': gd_flight.tail_no or '--',
            'std_utc': gd_flight.std_utc.strftime('%H:%M') if gd_flight.std_utc else '--',
            'sta_utc': gd_flight.sta_utc.strftime('%H:%M') if gd_flight.sta_utc else '--',
            'route': f"{gd_flight.origin_iata} → {gd_flight.destination_iata}",
            'crew_members': crew_data,
            'crew_count': len(crew_data),
            'processed_at': gd_flight.processed_at.strftime('%Y-%m-%d %H:%M:%S'),
            # ✨ NEW: Flitelink data
            'flitelink_status': gd_flight.flitelink_status,
            'flitelink_status_display': gd_flight.get_flitelink_status_display(),
            'flitelink_request_id': str(gd_flight.flitelink_request_id) if gd_flight.flitelink_request_id else None,
            'flitelink_submitted_at': gd_flight.flitelink_submitted_at.strftime('%Y-%m-%d %H:%M:%S') if gd_flight.flitelink_submitted_at else None,
            'flitelink_completed_at': gd_flight.flitelink_completed_at.strftime('%Y-%m-%d %H:%M:%S') if gd_flight.flitelink_completed_at else None,
            'flitelink_error_message': gd_flight.flitelink_error_message,
            'flitelink_retry_count': gd_flight.flitelink_retry_count,
            'can_submit_to_flitelink': gd_flight.can_submit_to_flitelink,
            'api_logs': api_logs_data,
        }
        
        return JsonResponse(data)
        
    except JEPPESSENGDFlight.DoesNotExist:
        return JsonResponse({'error': 'Flight not found'}, status=404)
    except Exception as e:
        logger.error(f"Error in jeppessen_flight_details: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


# ================================================================
# ✨ NEW: FLITELINK ACTION ENDPOINTS
# ================================================================

def flitelink_submit_flight(request, flight_id):
    """
    Manually submit a GD flight to Flitelink.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        gd_flight = get_object_or_404(JEPPESSENGDFlight, id=flight_id)
        
        # Check if can submit
        if not gd_flight.can_submit_to_flitelink:
            missing = []
            if not gd_flight.origin_icao: missing.append("origin_icao")
            if not gd_flight.destination_icao: missing.append("destination_icao")
            if not gd_flight.std_utc: missing.append("std_utc")
            if not gd_flight.sta_utc: missing.append("sta_utc")
            
            return JsonResponse({
                'error': f'Cannot submit - Missing: {", ".join(missing)}'
            }, status=400)
        
        # Submit
        from .tasks import submit_flight_to_flitelink
        result = submit_flight_to_flitelink.delay(gd_flight.id)
        
        return JsonResponse({
            'success': True,
            'message': f'Flight {gd_flight.flight_no} queued for submission',
            'task_id': result.id,
        })
        
    except Exception as e:
        logger.error(f"Error submitting flight: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


def flitelink_retry_submission(request, flight_id):
    """
    Retry a failed Flitelink submission.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        gd_flight = get_object_or_404(JEPPESSENGDFlight, id=flight_id)
        
        # Check if failed
        if gd_flight.flitelink_status != 'FAILED':
            return JsonResponse({
                'error': f'Cannot retry - Status is {gd_flight.get_flitelink_status_display()}'
            }, status=400)
        
        # Check retry limit
        max_retries = 3
        if gd_flight.flitelink_retry_count >= max_retries:
            return JsonResponse({
                'error': f'Maximum retries ({max_retries}) exceeded'
            }, status=400)
        
        # Increment retry count and reset status
        gd_flight.flitelink_retry_count += 1
        gd_flight.flitelink_status = 'PENDING'
        gd_flight.flitelink_error_message = None
        gd_flight.save()
        
        # Submit
        from .tasks import submit_flight_to_flitelink
        result = submit_flight_to_flitelink.delay(gd_flight.id)
        
        return JsonResponse({
            'success': True,
            'message': f'Retry queued for flight {gd_flight.flight_no}',
            'task_id': result.id,
            'retry_count': gd_flight.flitelink_retry_count,
        })
        
    except Exception as e:
        logger.error(f"Error retrying submission: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


def flitelink_refresh_status(request, flight_id):
    """
    Manually refresh Flitelink status of a flight.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        gd_flight = get_object_or_404(JEPPESSENGDFlight, id=flight_id)
        
        # Check if queued
        if gd_flight.flitelink_status != 'QUEUED':
            return JsonResponse({
                'error': f'Cannot check status - Status is {gd_flight.get_flitelink_status_display()}'
            }, status=400)
        
        # Check status
        from .tasks import check_submission_status
        success = check_submission_status(gd_flight)
        
        if success:
            gd_flight.refresh_from_db()
            return JsonResponse({
                'success': True,
                'status': gd_flight.flitelink_status,
                'status_display': gd_flight.get_flitelink_status_display(),
                'message': f'Status updated to {gd_flight.get_flitelink_status_display()}',
            })
        else:
            return JsonResponse({
                'error': 'Failed to check status'
            }, status=500)
        
    except Exception as e:
        logger.error(f"Error refreshing status: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


def flitelink_bulk_submit(request):
    """
    Bulk submit all unsubmitted flights from last 7 days.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        start_date = date.today() - timedelta(days=7)
        
        # Get unsubmitted flights
        unsubmitted = JEPPESSENGDFlight.objects.filter(
            flight_date__gte=start_date,
            flitelink_status='NOT_SUBMITTED'
        )
        
        # Filter to only submittable flights
        submittable = [f for f in unsubmitted if f.can_submit_to_flitelink]
        
        if not submittable:
            return JsonResponse({
                'success': True,
                'message': 'No flights to submit',
                'count': 0,
            })
        
        # Submit all
        from .tasks import submit_flight_to_flitelink
        
        submitted = 0
        for flight in submittable:
            try:
                submit_flight_to_flitelink.delay(flight.id)
                submitted += 1
            except Exception as e:
                logger.error(f"Error submitting flight {flight.id}: {e}")
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'Queued {submitted} flights for submission',
            'count': submitted,
        })
        
    except Exception as e:
        logger.error(f"Error in bulk submit: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)
    





# =================================================================================

# CREW TRANSPORTATION SYSTEM
 
# ================================================================================== 
"""
Django views for Crew Pick-Up Report Processor
"""

from django.shortcuts import render
from django.http import FileResponse, JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
import os
import tempfile
from werkzeug.utils import secure_filename
import pandas as pd
import sys
from datetime import datetime
from lxml import etree
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re


ALLOWED_EXTENSIONS = ['.xls', '.xlsx', '.xml', '.csv']
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB


class CrewReportProcessor:
    def __init__(self, input_file, output_dir='outputs'):
        """Initialize the processor with input file and output directory"""
        self.input_file = input_file
        self.output_dir = output_dir
        
        # Create output directory if it doesn't exist
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        
        # Position codes that indicate crew members (in priority order)
        self.position_codes = ['CP', 'FO', 'FP', 'SA', 'FA', 'MX', 'CC', 'CA']
        
        # Position priority for sorting
        self.position_priority = {pos: i for i, pos in enumerate(self.position_codes)}
    
    def parse_xml_excel(self, file_path):
        """Parse XML-formatted Excel file"""
        try:
            # Read file content
            with open(file_path, 'r', encoding='iso-8859-1') as f:
                content = f.read()
            
            # Parse XML
            parser = etree.XMLParser(encoding='iso-8859-1')
            tree = etree.fromstring(content.encode('iso-8859-1'), parser=parser)
            
            # Extract all rows
            rows = tree.xpath('//ss:Row', namespaces={'ss': 'urn:schemas-microsoft-com:office:spreadsheet'})
            
            data = []
            for row in rows:
                cells = row.xpath('.//ss:Data', namespaces={'ss': 'urn:schemas-microsoft-com:office:spreadsheet'})
                row_data = []
                
                # Handle sparse data (Excel XML may skip empty cells)
                current_index = 0
                for cell_elem in row.xpath('.//ss:Cell', namespaces={'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}):
                    # Check if cell has an index attribute
                    if '{urn:schemas-microsoft-com:office:spreadsheet}Index' in cell_elem.attrib:
                        target_index = int(cell_elem.attrib['{urn:schemas-microsoft-com:office:spreadsheet}Index']) - 1
                        # Fill empty cells up to target index
                        while current_index < target_index:
                            row_data.append('')
                            current_index += 1
                    
                    # Get cell data
                    data_elem = cell_elem.xpath('.//ss:Data', namespaces={'ss': 'urn:schemas-microsoft-com:office:spreadsheet'})
                    if data_elem:
                        row_data.append(data_elem[0].text.strip() if data_elem[0].text else '')
                    else:
                        row_data.append('')
                    current_index += 1
                
                # Ensure we have at least 10 columns
                while len(row_data) < 10:
                    row_data.append('')
                
                data.append(row_data)
            
            return data
            
        except Exception as e:
            print(f"Error parsing XML: {e}")
            # Try reading as regular Excel
            return self.parse_regular_excel(file_path)
    
    def parse_regular_excel(self, file_path):
        """Parse regular Excel file using pandas"""
        try:
            # Try reading with pandas
            df = pd.read_excel(file_path, header=None)
            return df.values.tolist()
        except Exception as e:
            print(f"Error reading Excel: {e}")
            return []
    
    def extract_date_from_data(self, data):
        """Extract date from the data"""
        for row in data:
            for cell in row:
                if isinstance(cell, str):
                    # Look for date patterns DD/MM/YYYY
                    date_match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', cell)
                    if date_match:
                        day = date_match.group(1).zfill(2)
                        month = date_match.group(2).zfill(2)
                        year = date_match.group(3)
                        return f"{day}-{month}-{year}"
        
        # If no date found, use current date
        return datetime.now().strftime("%d-%m-%Y")
    
    def find_header_row(self, data):
        """Find the header row index"""
        for i, row in enumerate(data):
            # Check if this row contains expected headers
            if any('DATE' in str(cell).upper() for cell in row[:2]):
                return i
            # Also check for 'Pos' column which is distinctive
            if 'Pos' in row or 'POS' in [str(cell).upper() for cell in row]:
                return i
        return None
    
    def extract_flight_number(self, flight_details):
        """Extract flight number from flight details string"""
        if not flight_details:
            return None
        
        # Extract flight number (e.g., WB601 from "WB601 dep 06:40")
        match = re.match(r'([A-Z0-9]+)', flight_details.strip())
        if match:
            return match.group(1)
        return None
    
    def process_data(self, data):
        """Process the data according to specifications"""
        # Find header row
        header_idx = self.find_header_row(data)
        if header_idx is None:
            print("Warning: Could not find header row, using row 4")
            header_idx = 4
        
        # Determine the column structure based on the number of columns
        if len(data[header_idx]) >= 10:
            # Original 10-column format
            time_idx = 1
            flight_idx = 4
            pos_idx = 6
            name_idx = 8
        else:
            # Already processed 4-column format
            time_idx = 0
            flight_idx = 1
            pos_idx = 2
            name_idx = 3
        
        # First, collect all crew data
        crew_data = []
        current_time = ''
        current_flight_details = ''
        current_flight_number = None
        
        for i in range(header_idx + 1, len(data)):
            row = data[i]
            
            # Ensure row has enough columns
            min_cols = 10 if len(data[header_idx]) >= 10 else 4
            while len(row) < min_cols:
                row.append('')
            
            # Get values
            time = str(row[time_idx]).strip() if row[time_idx] and str(row[time_idx]).strip() != 'nan' else ''
            flight_details = str(row[flight_idx]).strip() if row[flight_idx] and str(row[flight_idx]).strip() != 'nan' else ''
            pos = str(row[pos_idx]).strip() if row[pos_idx] and str(row[pos_idx]).strip() != 'nan' else ''
            name = str(row[name_idx]).strip() if row[name_idx] and str(row[name_idx]).strip() != 'nan' else ''
            
            # Skip duplicate header rows
            if (time.lower() == 'time' or 
                flight_details.lower() == 'flight details' or
                pos.lower() == 'pos' or
                name.lower() == 'name'):
                continue
            
            # Update current time and flight if this row has them
            if time and ':' in time:
                current_time = time
            if flight_details:
                current_flight_details = flight_details
                current_flight_number = self.extract_flight_number(flight_details)
            
            # Check if this row has position code
            if pos in self.position_codes and name:
                crew_data.append({
                    'time': current_time,
                    'flight_details': current_flight_details,
                    'flight_number': current_flight_number,
                    'pos': pos,
                    'name': name,
                    'position_priority': self.position_priority.get(pos, 999)
                })
        
        # Group crew by flight number
        flights_dict = {}
        for crew in crew_data:
            flight_num = crew['flight_number']
            if flight_num:
                if flight_num not in flights_dict:
                    flights_dict[flight_num] = []
                flights_dict[flight_num].append(crew)
        
        # Sort crew within each flight by position priority
        for flight_num in flights_dict:
            flights_dict[flight_num].sort(key=lambda x: x['position_priority'])
        
        # Build final data
        processed_data = []
        
        # Add header row
        processed_data.append(['Time', 'Flight Details', 'Pos', 'NAME'])
        
        # Add flights in order
        flight_order = sorted(flights_dict.keys(), 
                            key=lambda x: flights_dict[x][0]['time'] if flights_dict[x] else '')
        
        for i, flight_num in enumerate(flight_order):
            crew_list = flights_dict[flight_num]
            
            # Add all crew for this flight
            for crew in crew_list:
                processed_data.append([
                    crew['time'],
                    crew['flight_details'],
                    crew['pos'],
                    crew['name']
                ])
            
            # Add separator row between flights
            if i < len(flight_order) - 1:
                processed_data.append(['', '', '', ''])
        
        return processed_data
    
    def create_excel_report(self, data, report_date):
        """Create formatted Excel report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Crew Report"
        
        # Define sky blue fill for colored rows
        blue_fill = PatternFill(start_color="89cdff", end_color="89cdff", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Add colored empty row as FIRST row
        ws.append(['', '', '', ''])
        for cell in ws[1]:
            cell.fill = blue_fill
            cell.border = thin_border
            cell.font = Font(name='Times New Roman', size=16)
        
        # 2. Add title in row 2
        ws.merge_cells('A2:D2')
        title_cell = ws['A2']
        title_cell.value = f'Crew Pick-Up Report {report_date}'
        title_cell.font = Font(name='Times New Roman', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Process data to show Time and Flight Details only on first row of each flight
        processed_rows = []
        current_flight = None
        last_was_separator = False
        
        for i, row_data in enumerate(data):
            if i == 0:  # Header row
                processed_rows.append(row_data)
                last_was_separator = False
                continue
            
            # Skip duplicate header rows
            is_header = (str(row_data[0]).strip().lower() == 'time' and 
                        str(row_data[1]).strip().lower() == 'flight details' and
                        str(row_data[2]).strip().lower() == 'pos' and
                        str(row_data[3]).strip().lower() == 'name')
            if is_header:
                continue
            
            # Check if this is a separator row
            if all(not cell or str(cell).strip() == '' for cell in row_data):
                if not last_was_separator:
                    processed_rows.append(row_data)
                    last_was_separator = True
                current_flight = None
                continue
            
            last_was_separator = False
            
            # Get flight details
            time_val = str(row_data[0]).strip() if row_data[0] else ''
            flight_val = str(row_data[1]).strip() if row_data[1] else ''
            pos_val = str(row_data[2]).strip() if row_data[2] else ''
            name_val = str(row_data[3]).strip() if row_data[3] else ''
            
            # Check if this is a new flight
            if flight_val and flight_val != current_flight:
                current_flight = flight_val
                processed_rows.append([time_val, flight_val, pos_val, name_val])
            else:
                processed_rows.append(['', '', pos_val, name_val])
        
        # Add all processed rows to worksheet (starting from row 3)
        for row_data in processed_rows:
            ws.append(row_data)
        
        # 3. Format header row (row 3)
        for cell in ws[3]:
            cell.font = Font(name='Times New Roman', size=16, bold=True)
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 4. Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=4), start=4):
            is_separator = all(cell.value is None or str(cell.value).strip() == '' for cell in row)
            
            for col_idx, cell in enumerate(row, start=1):
                if is_separator:
                    cell.fill = blue_fill
                    cell.border = thin_border
                    cell.font = Font(name='Times New Roman', size=16)
                else:
                    cell.font = Font(name='Times New Roman', size=16)
                    
                    if col_idx == 1 and cell.value:
                        cell.font = Font(name='Times New Roman', size=16, bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx == 3:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    cell.border = thin_border
        
        # 5. Add colored empty row after last data row
        ws.append(['', '', '', ''])
        colored_row = ws.max_row
        for cell in ws[colored_row]:
            cell.fill = blue_fill
            cell.border = thin_border
            cell.font = Font(name='Times New Roman', size=16)
        
        # 6. Add warning message
        ws.append(['DUE TO OPERATIONAL REASONS THIS PICK-UP IS SUBJECT TO CHANGE!!!', '', '', ''])
        warning_row = ws.max_row
        ws.merge_cells(f'A{warning_row}:D{warning_row}')
        warning_cell = ws[f'A{warning_row}']
        warning_cell.font = Font(name='Times New Roman', size=16, bold=True)
        warning_cell.alignment = Alignment(horizontal='center', vertical='center')
        warning_cell.border = thin_border
        
        # 7. Add final colored empty row
        ws.append(['', '', '', ''])
        final_colored_row = ws.max_row
        for cell in ws[final_colored_row]:
            cell.fill = blue_fill
            cell.border = thin_border
            cell.font = Font(name='Times New Roman', size=16)
        
        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 60
        
        # Generate filename
        filename = f'Crew Pick-Up Report {report_date}.xlsx'
        filepath = os.path.join(self.output_dir, filename)
        
        # Save file
        wb.save(filepath)
        
        return filepath
    
    def process(self):
        """Main processing function"""
        print(f"Processing file: {self.input_file}")
        
        # Check if input file exists
        if not os.path.exists(self.input_file):
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        
        # Parse the input file
        print("Parsing input file...")
        if self.input_file.lower().endswith('.xls') or '<?xml' in open(self.input_file, 'r', encoding='iso-8859-1', errors='ignore').read(100):
            data = self.parse_xml_excel(self.input_file)
        else:
            data = self.parse_regular_excel(self.input_file)
        
        if not data:
            raise ValueError("No data found in input file")
        
        print(f"Found {len(data)} rows of data")
        
        # Extract date
        report_date = self.extract_date_from_data(data)
        print(f"Report date: {report_date}")
        
        # Process data
        print("Processing data...")
        processed_data = self.process_data(data)
        print(f"Processed {len(processed_data)} rows")
        
        # Create Excel report
        print("Creating Excel report...")
        output_file = self.create_excel_report(processed_data, report_date)
        
        print(f"Report generated successfully: {output_file}")
        return output_file


def crew_transportation_index(request):
    """Display the upload form"""
    return render(request, 'aimsintegration/transportation_index.html')


@require_http_methods(["POST"])
def upload_transport_file(request):
    """Handle file upload and processing"""
    temp_input = None
    temp_output_dir = None
    
    try:
        # Check if file was uploaded
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        file = request.FILES['file']
        
        # Check if file was selected
        if not file.name:
            return JsonResponse({'error': 'No file selected'}, status=400)
        
        # Check file size
        if file.size > MAX_FILE_SIZE:
            return JsonResponse({'error': 'File size exceeds 16MB limit'}, status=400)
        
        # Validate file extension
        filename = secure_filename(file.name)
        file_ext = os.path.splitext(filename)[1].lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            return JsonResponse({
                'error': f'Invalid file type. Allowed types: {", ".join(ALLOWED_EXTENSIONS)}'
            }, status=400)
        
        # Save uploaded file to temp directory
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp_file:
            for chunk in file.chunks():
                tmp_file.write(chunk)
            temp_input = tmp_file.name
        
        # Create temp output directory
        temp_output_dir = tempfile.mkdtemp()
        
        # Process the file
        processor = CrewReportProcessor(temp_input, temp_output_dir)
        output_file = processor.process()
        
        # Read the entire file into memory
        with open(output_file, 'rb') as f:
            file_content = f.read()
        
        # Create response with the file content
        from django.http import HttpResponse
        response = HttpResponse(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(output_file)}"'
        
        # Clean up temp files immediately (safe now that file is in memory)
        try:
            if temp_input and os.path.exists(temp_input):
                os.unlink(temp_input)
        except Exception as e:
            print(f"Error cleaning up temp input: {e}")
        
        try:
            import shutil
            if temp_output_dir and os.path.exists(temp_output_dir):
                shutil.rmtree(temp_output_dir)
        except Exception as e:
            print(f"Error cleaning up temp output: {e}")
        
        return response
        
    except Exception as e:
        print(f"Upload error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Clean up on error
        try:
            if temp_input and os.path.exists(temp_input):
                os.unlink(temp_input)
        except Exception as e:
            print(f"Error cleaning up temp input: {e}")
        
        try:
            import shutil
            if temp_output_dir and os.path.exists(temp_output_dir):
                shutil.rmtree(temp_output_dir)
        except Exception as e:
            print(f"Error cleaning up temp output: {e}")
        
        return JsonResponse({'error': str(e)}, status=500)

def get_directory_size(path):
    """Get total size of a directory and all its contents in bytes"""
    total_size = 0
    
    # Check if it's a file
    if os.path.isfile(path):
        return os.path.getsize(path)
    
    # If it's a directory, walk through it
    for dirpath, dirnames, filenames in os.walk(path):
        for filename in filenames:
            filepath = os.path.join(dirpath, filename)
            # Skip if it's a symbolic link to avoid counting twice
            if not os.path.islink(filepath):
                try:
                    total_size += os.path.getsize(filepath)
                except (OSError, FileNotFoundError):
                    # Handle permission errors or files that disappeared
                    pass
    
    return total_size

def format_time_since(backup):
    """Return human-readable time since backup."""
    if not backup:
        return "N/A"

    # Convert date to datetime for comparison
    if isinstance(backup.backup_date, date) and not isinstance(backup.backup_date, datetime):
        backup_datetime = datetime.combine(backup.backup_date, datetime.min.time())
        backup_datetime = timezone.make_aware(backup_datetime, timezone.get_current_timezone())
    else:
        backup_datetime = backup.backup_date

    delta = timezone.now() - backup_datetime

    if delta.days >= 1:
        return f"{delta.days} Days"

    hours = delta.seconds // 3600
    if hours >= 1:
        return f"{hours} Hours"

    minutes = delta.seconds // 60
    if minutes >= 1:
        return f"{minutes} Minutes"

    return "Just now"


def backup_view(request):
    # Path to your static folder's subdirectory
    backup_type = request.GET.get("type", "monthly").lower().strip()
    folder_path = os.path.join(settings.BACKUP_CREW_DOCUMENTS_PATH, backup_type)

    # Get all files in the folder
    files = []
    total_size = 0

    # backups = Backup.objects.filter(backup_type=backup_type, status="success")

    # for backup in backups:
    #     files.append({
    #         'name': backup.name,
    #         # Static URL to access the file from template
    #         'url': f'/media/crew_documents/{backup_type}/{backup.name}',
    #         'log': f'Run: {backup.start_time.strftime("%Y-%m-%d %H:%M:%S")}, for {backup.duration_minutes} minutes',
    #         'size': format_file_size(backup.backup_size),
    #         'date': backup.backup_date.strftime('%Y-%m-%d')
    #     })
    #     total_size += backup.backup_size or 0

    if os.path.exists(folder_path):
        all_items = os.listdir(folder_path)
        for filename in all_items:
            file_path = os.path.join(folder_path, filename)
            files.append({
                'name': filename,
                'url': f'/media/crew_documents/{backup_type}/{filename}',
                'size': '' if os.path.isdir(file_path) else format_file_size(get_directory_size(file_path)),
                'date': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d')
            })
            total_size += get_directory_size(file_path)

    last_successful_monthly_backup = Backup.objects.filter(backup_type="monthly", status="success").order_by('-backup_date').first()
    last_successful_weekly_backup = Backup.objects.filter(backup_type="weekly", status="success").order_by('-backup_date').first()
    time_since_monthly_backup = format_time_since(last_successful_monthly_backup)
    time_since_weekly_backup = format_time_since(last_successful_weekly_backup)

    context = {
        'files': files,
        'backup_size': format_file_size(total_size),
        "backup_type": backup_type,
        "time_since_monthly_backup": time_since_monthly_backup,
        "time_since_weekly_backup": time_since_weekly_backup,
    }
    return render(request, 'aimsintegration/backup.html', context)

def backup_folder_view(request, folder_name):
    # Path to your static folder's subdirectory
    backup_type = request.GET.get("type", "").lower().strip()
    folders = folder_name.split("&&")
    folder_path = os.path.join(settings.BASE_DIR, 'media', 'crew_documents', backup_type, *folders)

    # Get all files in the folder
    files = []
    searched = []

    q = request.GET.get("q", "").lower().strip()

    if os.path.exists(folder_path):
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            # if os.path.isfile(file_path):
            item = {
                'name': filename,
                # Static URL to access the file from template
                'url': '/'.join(['media', 'crew_documents', backup_type, *folders, filename]),
                'extension': filename.split('.')[-1] if '.' in filename else 'folder',
                'size': '' if os.path.isdir(file_path) else format_file_size(get_directory_size(file_path)),
                'date': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d')
            }
            files.append(item)

            # if q:
            #     if q in filename.lower():
            #         searched.append(item)
            
            if q:
                for dirpath, dirnames, filenames in os.walk(file_path):
                    for flname in filenames:
                        if q in flname.lower():
                            full_path = os.path.join(dirpath, flname)
                            searched.append({
                                'name': flname,
                                # Static URL to access the file from template
                                'url': '/'.join(['','media', 'crew_documents', backup_type, *folders, os.path.relpath(full_path, folder_path)]),
                                'extension': flname.split('.')[-1] if '.' in flname else 'folder',
                                'size': '' if os.path.isdir(full_path) else format_file_size(get_directory_size(full_path)),
                                'date': datetime.fromtimestamp(os.path.getmtime(full_path)).strftime('%Y-%m-%d')
                            })

    context = {
        'files': searched if q else files,
        'folder_name': folder_name,
        'display_folder_name': folder_name.replace("&&", " > "),
        'q': q or '',
        "backup_type": backup_type
    }
    return render(request, 'aimsintegration/backup_folder_view.html', context)


from django.views.decorators.http import require_http_methods
@require_http_methods(["GET"])
def preview_file(request, file_path):
    """
    Returns file preview based on file type
    """
    try:
        # Ensure the file exists and is within your allowed directory
        full_path = Path(file_path)
        
        if not full_path.exists():
            return JsonResponse({'error': 'File not found'}, status=404)
        
        file_extension = full_path.suffix.lower()
        
        # For PDFs, return the file directly
        if file_extension == '.pdf':
            return FileResponse(
                open(full_path, 'rb'),
                content_type='application/pdf'
            )
        
        # For other files, return file info
        return JsonResponse({
            'file_name': full_path.name,
            'file_type': file_extension,
            'file_url': request.build_absolute_uri(f'/media/documents/{full_path.name}')
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from io import BytesIO
import zipfile

def download_folder_zip(request, folder_name):
    """Download all files in a folder as a ZIP file (or only searched files if q parameter exists)"""
    backup_type = request.GET.get("type", "").lower().strip()
    folders = folder_name.split("&&")
    folder_path = os.path.join(settings.BASE_DIR, 'media', 'crew_documents', backup_type, *folders)

    print("folder_path:", folder_path, " folders:", folders)
    if not os.path.exists(folder_path):
        raise Http404("Folder not found")
    
    # Get search query
    q = request.GET.get("q", "").lower().strip()

    print(" q: ", q)
    
    # Create a BytesIO buffer to hold the ZIP file in memory
    buffer = BytesIO()

    wrap_inside_folder = (backup_type == "archive")
    zip_root = folder_name if wrap_inside_folder else ""
    
    # Create a ZIP file
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        if q:
            # When there's a search query, only include matching files
            for filename in os.listdir(folder_path):
                file_path = os.path.join(folder_path, filename)
                
                # Walk through the directory tree
                for dirpath, dirnames, filenames in os.walk(file_path):
                    for flname in filenames:
                        # Only include files that match the search query
                        if q in flname.lower():
                            full_path = os.path.join(dirpath, flname)
                            # Calculate the archive name (relative path from folder_path)
                            arcname = os.path.relpath(full_path, folder_path)
                            zip_file.write(full_path, arcname)
        else:
            # No search query - download all files in the folder
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    relative_path = os.path.relpath(file_path, folder_path)
                    
                    # Calculate the archive name (relative path from folder_path)
                    # Wrap inside root folder if needed
                    if wrap_inside_folder:
                        arcname = os.path.join(zip_root, relative_path)
                    else:
                        arcname = relative_path
                    print("file_path:", file_path, " archive_name:", arcname)
                    zip_file.write(file_path, arcname)
    
    # Reset buffer position to the beginning
    buffer.seek(0)
    
    # Create the HTTP response with ZIP file
    response = HttpResponse(buffer.getvalue(), content_type='application/zip')
    
    # Set the filename based on whether it's a search or full folder download
    if q:
        zip_filename = f"{folders[-1]}_{q}.zip" if folders else f"{q}.zip"
    else:
        zip_filename = f"{folders[-1]}.zip" if folders else "backup.zip"
    
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    
    return response

def download_one_file(request, file_path):
    """Download a single file"""
    full_path = os.path.join(settings.BASE_DIR, file_path)

    if not os.path.exists(full_path):
        raise Http404("File not found")

    with open(full_path, 'rb') as f:
        file_data = f.read()

    response = HttpResponse(file_data, content_type='application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(full_path)}"'
    return response

import os
import shutil
from django.http import JsonResponse, Http404
from django.conf import settings


def archive_crew_documents_by_wb(request, wb_number):
    """
    Archive crew documents that match a given wb_number.
    Instead of zipping, we copy matching files to:
    media/crew_documents/archive/<wb_number>/
    """

    source_root = os.path.join(
        settings.BASE_DIR, 
        "media", 
        "crew_documents", 
        "weekly", 
        "Crew Documents"
    )

    if not os.path.exists(source_root):
        raise Http404("Source folder not found")

    # wb_number is the search text
    search_term = wb_number.lower().strip()

    # Destination archive folder
    archive_root = os.path.join(
        settings.BASE_DIR,
        "media",
        "crew_documents",
        "archive",
        wb_number
    )

    # Ensure destination exists
    os.makedirs(archive_root, exist_ok=True)

    # Track copied files for reporting
    copied_files = []

    # Walk through documents folder
    for dirpath, dirnames, filenames in os.walk(source_root):
        for filename in filenames:
            if search_term in filename.lower():

                src_file = os.path.join(dirpath, filename)

                # Keep relative directory structure
                relative_path = os.path.relpath(dirpath, source_root)
                dest_dir = os.path.join(archive_root, relative_path)

                os.makedirs(dest_dir, exist_ok=True)

                dest_file = os.path.join(dest_dir, filename)
                shutil.copy2(src_file, dest_file)

                copied_files.append(dest_file)

    return JsonResponse({
        "status": "success",
        "archived_to": archive_root,
        "files_copied": copied_files,
        "count": len(copied_files),
    })

def archive_view(request):
    left_crew_members = []

    # Get 10 records of crews, sorted by date of leaving
    left_crew_members = []

    # Get 10 records of crews, sorted by date of leaving
    query = request.GET.get('query', '').strip()
    sort = request.GET.get('sort', 'desc').strip()
    sorting = '-date_of_leaving' if sort == 'desc' else 'date_of_leaving'
    queryset = CrewDocumentsArchive.objects.order_by(sorting)

    # if there is query, return all the matched records, ignore pagination
    if query:
        queryset = CrewDocumentsArchive.objects.filter(
            Q(wb_number__icontains=query) |
            Q(crew_name__icontains=query) |
            Q(position__icontains=query)
        ).order_by(sorting)

    paginator = Paginator(queryset, 10)  # 10 per page

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
   
    # add 24 months on exit date and put it on a field called, archive date
    for record in list(page_obj.object_list.values()):
        left_crew_members.append({
            'id': record['wb_number'],
            'wb_number': f"WB{int(record['wb_number']):04d}",
            'crew_name': record['crew_name'],
            'exit_date': record['date_of_leaving'].strftime('%Y-%m-%d'),
            'position': record['position'],
            'archive_date': record['archive_date'].strftime('%Y-%m-%d') if record['archive_date'] else '--',
            'archived': record['archived'],
            'archive_path': record['archive_path'],
        })

    context = {
        "page": page_obj.number,
        "pages": paginator.num_pages,
        "left_crew_members": left_crew_members,
        "query": query or '',
        "sort": sort,
    }
    return render(request, 'aimsintegration/archive.html', context)

def archive_folder_view(request):
    folder_path = os.path.join(settings.BACKUP_CREW_DOCUMENTS_PATH, 'archive')
    files=[]

    if os.path.exists(folder_path):
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            files.append({
                'name': filename,
                'url': f'/media/crew_documents/archive/{filename}',
                'size': '' if os.path.isdir(file_path) else format_file_size(get_directory_size(file_path)),
                'extension': filename.split('.')[-1] if '.' in filename else 'folder',
                'date': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d')
            })
            # total_size += get_directory_size(file_path)

    context = {
        "files": files
    }
    return render(request, 'aimsintegration/archives_view.html', context)

def archive_single_folder_view(request, folder_name):
    folders = folder_name.split("&&")
    folder_path = os.path.join(settings.BACKUP_CREW_DOCUMENTS_PATH, 'archive', *folders)
    files=[]

    wb_number = folders[-1] if len(folders) > 0 else ''
    is_number = wb_number.isdigit()
    if is_number:
        crew = CrewDocumentsArchive.objects.filter(wb_number=wb_number).first()
    else:
        crew = None

    if os.path.exists(folder_path):
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            files.append({
                'name': filename,
                'url': '/'.join(['media', 'crew_documents', 'archive', *folders, filename]),
                'size': '' if os.path.isdir(file_path) else format_file_size(get_directory_size(file_path)),
                'extension': filename.split('.')[-1] if '.' in filename else 'folder',
                'date': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d')
            })
            # total_size += get_directory_size(file_path)

    context = {
        "files": files,
        'folder_name': folder_name,
        'crew': crew.crew_name if crew else 'Initial Archive',
        'archive_date': crew.archive_date.strftime('%Y-%m-%d') if crew else datetime.fromtimestamp(os.path.getmtime(folder_path)).strftime('%Y-%m-%d'),
    }
    return render(request, 'aimsintegration/archives_folder_view.html', context)
